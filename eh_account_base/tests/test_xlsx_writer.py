# -*- encoding: utf-8 -*-
##############################################################################
#
# ERP Heritage
# Copyright (C) 2026 (https://www.erpheritage.com.au/)
#
##############################################################################
"""
Unit tests for XlsxReportWriter.

These tests do not require Odoo's DB, only openpyxl. They construct a
synthetic payload and verify that:

* The writer returns bytes that openpyxl can re open without errors.
* Title row, meta rows, header row, and data rows land in expected positions.
* Numeric cells get the correct number format and right alignment.
* Totals row carries bold font and a top border.
* Empty payloads do not crash.
* The sheet name is truncated to Excel's 31 character limit.
"""

import io

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from odoo.tests import tagged

from odoo.addons.eh_account_base.tools.xlsx_writer import XlsxReportWriter
from .common import EhAccountUnitTestCase


PAYLOAD = {
    'columns': [
        {'expression_label': 'account', 'name': 'Account',
         'figure_type': 'string'},
        {'expression_label': 'opening_debit', 'name': 'Opening DB',
         'figure_type': 'monetary'},
        {'expression_label': 'opening_credit', 'name': 'Opening CR',
         'figure_type': 'monetary'},
        {'expression_label': 'closing_debit', 'name': 'Closing DB',
         'figure_type': 'monetary'},
        {'expression_label': 'closing_credit', 'name': 'Closing CR',
         'figure_type': 'monetary'},
    ],
    'lines': [
        {
            'id': 'account-1',
            'name': '1000 Cash on Hand',
            'level': 1,
            'columns': [
                {'expression_label': 'opening_debit', 'value': 1000.0},
                {'expression_label': 'opening_credit', 'value': 0.0},
                {'expression_label': 'closing_debit', 'value': 1500.0},
                {'expression_label': 'closing_credit', 'value': 0.0},
            ],
        },
        {
            'id': 'account-2',
            'name': '4000 Sales Revenue',
            'level': 1,
            'columns': [
                {'expression_label': 'opening_debit', 'value': 0.0},
                {'expression_label': 'opening_credit', 'value': 1000.0},
                {'expression_label': 'closing_debit', 'value': 0.0},
                {'expression_label': 'closing_credit', 'value': 1500.0},
            ],
        },
    ],
    'totals': {
        'opening_debit': 1000.0,
        'opening_credit': 1000.0,
        'closing_debit': 1500.0,
        'closing_credit': 1500.0,
    },
    'meta': {
        'date_from': '2026-01-01',
        'date_to': '2026-12-31',
        'posted_only': True,
    },
    'generated_at': '2026-04-30T12:00:00',
}


@tagged('eh_account_base', 'unit')
class TestXlsxWriterShape(EhAccountUnitTestCase):

    def _render(self, payload=None, name="Trial Balance"):
        content = XlsxReportWriter(name).write_payload(payload or PAYLOAD)
        return content, load_workbook(io.BytesIO(content))

    def test_returns_xlsx_bytes(self):
        content, wb = self._render()
        self.assertIsInstance(content, bytes)
        # XLSX is a ZIP archive; magic bytes are 'PK'.
        self.assertEqual(content[:2], b'PK')
        self.assertEqual(wb.active.title, 'Report')

    def test_title_row_present(self):
        _, wb = self._render(name="My Custom Title")
        ws = wb.active
        self.assertEqual(ws.cell(row=1, column=1).value, "My Custom Title")
        self.assertTrue(ws.cell(row=1, column=1).font.bold)

    def test_meta_rows_show_period_and_filters(self):
        _, wb = self._render()
        ws = wb.active
        # Search the first 6 rows for the meta info.
        meta_text = " ".join(
            (ws.cell(row=r, column=1).value or '')
            for r in range(1, 7)
        )
        self.assertIn("2026-01-01", meta_text)
        self.assertIn("2026-12-31", meta_text)
        self.assertIn("Posted entries only", meta_text)

    def test_column_headers_appear_with_correct_text(self):
        _, wb = self._render()
        ws = wb.active
        header_row = self._find_row_starting_with(ws, "Account")
        self.assertIsNotNone(header_row)
        self.assertEqual(ws.cell(row=header_row, column=1).value, "Account")
        self.assertEqual(ws.cell(row=header_row, column=2).value, "Opening DB")
        self.assertEqual(ws.cell(row=header_row, column=5).value, "Closing CR")

    def test_data_lines_carry_correct_values(self):
        _, wb = self._render()
        ws = wb.active
        cash_row = self._find_row_starting_with(ws, "1000 Cash on Hand")
        self.assertIsNotNone(cash_row)
        self.assertEqual(ws.cell(row=cash_row, column=2).value, 1000.0)
        self.assertEqual(ws.cell(row=cash_row, column=4).value, 1500.0)

    def test_monetary_cells_have_accounting_format(self):
        _, wb = self._render()
        ws = wb.active
        cash_row = self._find_row_starting_with(ws, "1000 Cash on Hand")
        money_cell = ws.cell(row=cash_row, column=2)
        self.assertIn('#,##0.00', money_cell.number_format)
        self.assertEqual(money_cell.alignment.horizontal, 'right')

    def test_totals_row_present_with_bold_font_and_top_border(self):
        _, wb = self._render()
        ws = wb.active
        totals_row = self._find_row_starting_with(ws, "Totals")
        self.assertIsNotNone(totals_row)
        totals_cell = ws.cell(row=totals_row, column=2)
        self.assertEqual(totals_cell.value, 1000.0)
        self.assertTrue(totals_cell.font.bold)
        self.assertIsNotNone(totals_cell.border.top.style)

    def test_column_widths_set(self):
        _, wb = self._render()
        ws = wb.active
        # First column (name) is wider than monetary columns.
        first_width = ws.column_dimensions[get_column_letter(1)].width
        money_width = ws.column_dimensions[get_column_letter(2)].width
        self.assertGreater(first_width, money_width)

    def test_empty_payload_does_not_crash(self):
        content = XlsxReportWriter("Empty").write_payload({
            'columns': [],
            'lines': [],
            'totals': {},
            'generated_at': '',
        })
        self.assertIsInstance(content, bytes)

    def test_sheet_name_truncated_to_31_chars(self):
        very_long = "A" * 50
        writer = XlsxReportWriter("Report", sheet_name=very_long)
        # openpyxl raises if title exceeds 31 chars; survival is the assertion.
        content = writer.write_payload(PAYLOAD)
        wb = load_workbook(io.BytesIO(content))
        self.assertLessEqual(len(wb.active.title), 31)

    def test_negative_values_render(self):
        payload = {
            'columns': [
                {'expression_label': 'account', 'name': 'Account',
                 'figure_type': 'string'},
                {'expression_label': 'balance', 'name': 'Balance',
                 'figure_type': 'monetary'},
            ],
            'lines': [
                {'id': 'l1', 'name': 'Loss', 'level': 1,
                 'columns': [{'expression_label': 'balance', 'value': -1234.56}]},
            ],
            'totals': {'balance': -1234.56},
            'generated_at': '',
        }
        content, wb = self._render(payload=payload)
        ws = wb.active
        loss_row = self._find_row_starting_with(ws, "Loss")
        self.assertEqual(ws.cell(row=loss_row, column=2).value, -1234.56)

    @staticmethod
    def _find_row_starting_with(ws, text):
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == text:
                return r
        return None
