# -*- encoding: utf-8 -*-
##############################################################################
#
# ERP Heritage
# Copyright (C) 2026 (https://www.erpheritage.com.au/)
#
##############################################################################
"""
XLSX writer for dynamic report payloads.

Consumes the standard payload shape produced by every dynamic report
handler:

    {
        'columns': [
            {'expression_label': str, 'name': str, 'figure_type': str},
            ...
        ],
        'lines': [
            {
                'id': str, 'name': str, 'level': int,
                'columns': [{'expression_label': str, 'value': any}, ...],
                'unfoldable': bool,
                'meta': dict,
            },
            ...
        ],
        'totals': {expression_label: number, ...},
        'meta': {date_from, date_to, posted_only, ...},
        'generated_at': iso datetime,
    }

Convention: payload['columns'][0] is always the line label column, rendered
from line['name']. payload['columns'][1:] are the value columns, rendered
from line['columns'][i] where i is the value column index.

Output is a single sheet workbook with:

* A title row with the report name.
* Optional metadata rows (period, posted only flag, generated at).
* A header row with column titles.
* One data row per line.
* A totals row with a top border, when totals are present.

The writer is intentionally a plain Python class so unit tests do not need
the full Odoo registry. openpyxl is the only external dependency, and it is
already a standard Odoo requirement.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# Number formats for openpyxl. The accounting convention is to render
# negatives in red and positive numbers without a sign.
_FORMAT_MONETARY = '#,##0.00;[Red](#,##0.00)'
_FORMAT_INTEGER = '#,##0'
_FORMAT_FLOAT = '#,##0.0000'
_FORMAT_PERCENT = '0.00%'
_FORMAT_DATE = 'yyyy-mm-dd'
_FORMAT_DATETIME = 'yyyy-mm-dd hh:mm:ss'

_NUMERIC_FIGURE_TYPES = frozenset({'monetary', 'integer', 'float', 'percentage'})


class XlsxReportWriter:
    """Render a dynamic report payload to XLSX bytes.

    Use:

        writer = XlsxReportWriter("Trial Balance")
        content_bytes = writer.write_payload(payload)

    The writer is single use. Construct a new instance per render.
    """

    DEFAULT_FONT = Font(name='Arial', size=10)
    HEADER_FONT = Font(name='Arial', size=10, bold=True)
    TITLE_FONT = Font(name='Arial', size=14, bold=True)
    META_FONT = Font(name='Arial', size=9, italic=True, color='666666')
    TOTAL_FONT = Font(name='Arial', size=10, bold=True)
    HEADER_FILL = PatternFill(
        start_color='F0F0F0', end_color='F0F0F0', fill_type='solid',
    )
    TOP_BORDER = Border(top=Side(border_style='thin', color='000000'))

    def __init__(self, report_name, sheet_name='Report'):
        self.report_name = report_name or "Report"
        # Excel imposes a 31 character maximum on sheet names.
        self.sheet_name = (sheet_name or "Report")[:31]
        self.workbook = Workbook()
        self.ws = self.workbook.active
        self.ws.title = self.sheet_name
        self._row = 1
        self._currency = None

    BRAND_FONT = Font(
        name='Arial', size=8, italic=True, color='6C757D',
    )
    BRAND_FOOTER_TEXT = "Made with 🤍 from Melbourne by ERP Heritage"

    def write_payload(self, payload, meta=None):
        meta = meta if meta is not None else (payload.get('meta') or {})
        generated_at = payload.get('generated_at') or ''
        self._currency = payload.get('currency') or None

        self._write_title()
        self._write_meta(meta, generated_at)
        self._write_blank_row()

        columns = payload.get('columns') or []
        self._write_column_headers(columns)
        for line in (payload.get('lines') or []):
            self._write_line(line, columns)
        if payload.get('totals'):
            self._write_totals(payload['totals'], columns)

        self._set_column_widths(columns)
        self._write_brand_footer()
        return self._serialise()

    def _write_brand_footer(self):
        """Write the ERP Heritage Melbourne brand line as the last
        non-data row of the sheet. Single cell, italic, muted grey, so
        it does not compete with the report content above.
        """
        self._row += 1
        cell = self.ws.cell(
            row=self._row, column=1, value=self.BRAND_FOOTER_TEXT,
        )
        cell.font = self.BRAND_FONT
        cell.alignment = Alignment(horizontal='left')

    # ---- title and meta rows ----

    def _write_title(self):
        ws = self.ws
        cell = ws.cell(row=self._row, column=1, value=self.report_name)
        cell.font = self.TITLE_FONT
        self._row += 1

    def _write_meta(self, meta, generated_at):
        ws = self.ws
        details = []
        date_from = meta.get('date_from')
        date_to = meta.get('date_to')
        if date_from or date_to:
            details.append("Period: %s to %s" % (date_from or '', date_to or ''))
        if 'posted_only' in meta:
            details.append(
                "Posted entries only" if meta['posted_only']
                else "All entries (including draft)",
            )
        if 'show_zero' in meta and meta['show_zero']:
            details.append("Including zero balance accounts")
        if self._currency:
            if self._currency.get('multi_currency'):
                details.append("Multi-currency scope")
            elif self._currency.get('name'):
                sym = self._currency.get('symbol') or ''
                details.append(
                    "Currency: %s%s" % (
                        self._currency['name'],
                        " (%s)" % sym if sym else "",
                    ),
                )
        if details:
            ws.cell(
                row=self._row, column=1, value=" | ".join(details),
            ).font = self.META_FONT
            self._row += 1
        if generated_at:
            ws.cell(
                row=self._row, column=1,
                value="Generated: %s" % generated_at,
            ).font = self.META_FONT
            self._row += 1

    def _write_blank_row(self):
        self._row += 1

    # ---- column header row ----

    def _write_column_headers(self, columns):
        ws = self.ws
        for i, col_def in enumerate(columns, start=1):
            cell = ws.cell(
                row=self._row, column=i, value=col_def.get('name', ''),
            )
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            if col_def.get('figure_type') in _NUMERIC_FIGURE_TYPES:
                cell.alignment = Alignment(horizontal='right')
        self._row += 1

    # ---- data lines ----

    def _write_line(self, line, columns):
        ws = self.ws
        # Column 1: line name. Indent if this is a sub line.
        name_cell = ws.cell(
            row=self._row, column=1, value=line.get('name', ''),
        )
        name_cell.font = (
            self.TOTAL_FONT if line.get('level', 1) == 0 else self.DEFAULT_FONT
        )
        level = max(int(line.get('level', 1)) - 1, 0)
        if level:
            name_cell.alignment = Alignment(indent=level)

        # Subsequent columns: line.columns[i] -> sheet column i + 2.
        for i, line_col in enumerate(line.get('columns') or []):
            sheet_col = i + 2
            col_def = columns[i + 1] if i + 1 < len(columns) else {}
            value = line_col.get('value')
            cell = ws.cell(row=self._row, column=sheet_col, value=value)
            cell.font = self.DEFAULT_FONT
            self._apply_figure_type(cell, col_def.get('figure_type', 'string'))

        self._row += 1

    # ---- totals row ----

    def _write_totals(self, totals, columns):
        ws = self.ws
        label_cell = ws.cell(row=self._row, column=1, value="Totals")
        label_cell.font = self.TOTAL_FONT
        label_cell.border = self.TOP_BORDER

        for i, col_def in enumerate(columns[1:], start=2):
            label = col_def.get('expression_label')
            value = totals.get(label) if label else None
            cell = ws.cell(row=self._row, column=i, value=value)
            cell.font = self.TOTAL_FONT
            cell.border = self.TOP_BORDER
            self._apply_figure_type(cell, col_def.get('figure_type', 'string'))

        self._row += 1

    # ---- formatting and widths ----

    def _monetary_format(self):
        """Build the Excel number format string from the payload currency.

        When the payload has no currency block (legacy or multi_currency),
        fall back to the static format that just renders thousands and
        two decimals.
        """
        if not self._currency or self._currency.get('multi_currency'):
            return _FORMAT_MONETARY
        decimals = int(self._currency.get('decimal_places') or 2)
        symbol = self._currency.get('symbol') or ''
        symbol_lit = symbol.replace('"', '\\"')
        if decimals <= 0:
            digits = '0'
        else:
            digits = '0.' + ('0' * decimals)
        if not symbol_lit:
            return f'#,##{digits};[Red](#,##{digits})'
        if self._currency.get('position') == 'before':
            return (
                f'"{symbol_lit}" #,##{digits};'
                f'[Red]"{symbol_lit}" (#,##{digits})'
            )
        return (
            f'#,##{digits} "{symbol_lit}";'
            f'[Red](#,##{digits}) "{symbol_lit}"'
        )

    def _apply_figure_type(self, cell, figure_type):
        if figure_type == 'monetary':
            cell.number_format = self._monetary_format()
            cell.alignment = Alignment(horizontal='right')
        elif figure_type == 'integer':
            cell.number_format = _FORMAT_INTEGER
            cell.alignment = Alignment(horizontal='right')
        elif figure_type == 'float':
            cell.number_format = _FORMAT_FLOAT
            cell.alignment = Alignment(horizontal='right')
        elif figure_type == 'percentage':
            cell.number_format = _FORMAT_PERCENT
            cell.alignment = Alignment(horizontal='right')
        elif figure_type == 'date':
            cell.number_format = _FORMAT_DATE
        elif figure_type == 'datetime':
            cell.number_format = _FORMAT_DATETIME
        # 'string', 'boolean', and unknown types use the default.

    def _set_column_widths(self, columns):
        ws = self.ws
        # Name column gets extra width.
        ws.column_dimensions[get_column_letter(1)].width = 35
        for i, col_def in enumerate(columns[1:], start=2):
            figure_type = col_def.get('figure_type', 'string')
            width = self._width_for(figure_type)
            ws.column_dimensions[get_column_letter(i)].width = width

    @staticmethod
    def _width_for(figure_type):
        if figure_type in ('monetary', 'float'):
            return 16
        if figure_type == 'integer':
            return 12
        if figure_type == 'percentage':
            return 12
        if figure_type in ('date', 'datetime'):
            return 16
        return 22

    # ---- serialisation ----

    def _serialise(self):
        buf = io.BytesIO()
        self.workbook.save(buf)
        return buf.getvalue()
