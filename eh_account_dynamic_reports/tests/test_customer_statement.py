# -*- encoding: utf-8 -*-
##############################################################################
#
# ERP Heritage
# Copyright (C) 2026 (https://www.erpheritage.com.au/)
#
##############################################################################
"""Open-item customer statement: a period settlement of a prior invoice must
reduce the amount due, not leave a phantom balance (regression for the
overstatement where a clearing payment was excluded by the residual filter)."""

import io
from unittest.mock import patch

from openpyxl import load_workbook

from odoo.exceptions import AccessError
from odoo.tests import new_test_user, tagged

from odoo.addons.eh_account_base.tests.common import (
    EhAccountIntegrationTestCase,
)


@tagged('eh_account_dynamic_reports', 'integration', 'post_install',
        '-at_install')
class TestCustomerStatement(EhAccountIntegrationTestCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.handler = cls.env[
            'eh.account.dynamic.report.handler.customer_statement']
        cls.vendor_handler = cls.env[
            'eh.account.dynamic.report.handler.vendor_statement']
        cls.report = cls.env['eh.account.dynamic.report'].search(
            [('code', '=', 'customer_statement')], limit=1)
        if not cls.report:
            cls.report = cls.env['eh.account.dynamic.report'].create({
                'code': 'customer_statement',
                'name': 'Customer Statement',
                'handler_model': (
                    'eh.account.dynamic.report.handler.customer_statement'
                ),
            })
        cls.partner = cls.env['res.partner'].create({'name': 'Stmt Cust'})
        cls.income = cls._ensure_account(
            cls.env, '4001', 'Stmt Income', 'income')
        cls.sale_journal = cls.env['account.journal'].search(
            [('type', '=', 'sale'),
             ('company_id', '=', cls.env.company.id)], limit=1)
        if not cls.sale_journal:
            cls.sale_journal = cls.env['account.journal'].create({
                'name': 'Stmt Sales', 'code': 'STMTS', 'type': 'sale',
                'company_id': cls.env.company.id})
        cls.bank_journal = cls.env['account.journal'].search(
            [('type', '=', 'bank'),
             ('company_id', '=', cls.env.company.id)], limit=1)
        if not cls.bank_journal:
            cls.bank_journal = cls.env['account.journal'].create({
                'name': 'Stmt Bank', 'code': 'STMTB', 'type': 'bank',
                'company_id': cls.env.company.id})

    def _invoice(self, amount, date):
        inv = self.env['account.move'].create({
            'move_type': 'out_invoice', 'partner_id': self.partner.id,
            'invoice_date': date, 'journal_id': self.sale_journal.id,
            'invoice_line_ids': [(0, 0, {
                'name': 'x', 'quantity': 1, 'price_unit': amount,
                'account_id': self.income.id})]})
        inv.action_post()
        return inv

    def _pay(self, inv, date):
        self.env['account.payment.register'].with_context(
            active_model='account.move', active_ids=inv.ids).create({
                'payment_date': date,
                'journal_id': self.bank_journal.id,
            })._create_payments()

    def _statement_payload(self, date_from, date_to, stype='open_item'):
        return self.handler.compute({
            'partner_id': self.partner.id,
            'date': {'date_from': date_from, 'date_to': date_to},
            'company_ids': self.env.company.ids,
            'statement_type': stype,
        })

    def _amount_due(self, date_from, date_to, stype='open_item'):
        payload = self._statement_payload(date_from, date_to, stype=stype)
        return abs(payload['totals']['amount_due'])

    def _filtered_amount_due(
        self, handler, account_type_ids, stype='open_item',
        date_from='2026-01-01', date_to='2026-12-31',
    ):
        payload = handler.compute({
            'partner_id': self.partner.id,
            'date': {'date_from': date_from, 'date_to': date_to},
            'company_ids': self.env.company.ids,
            'posted_only': True,
            'statement_type': stype,
            'account_type_ids': account_type_ids,
        })
        return abs(payload['totals']['amount_due'])

    def test_kwd_precision_keeps_one_fils_statement_balance(self):
        kwd = self.env['res.currency'].with_context(active_test=False).search(
            [('name', '=', 'KWD')], limit=1,
        )
        self.assertTrue(kwd)
        lines, total_due = self.handler._build_lines(
            0.001, [], '2026-06-30', statement_type='activity',
            currency=kwd,
        )
        opening = next(line for line in lines if line['id'] == 'opening')
        closing = next(
            line for line in lines if line['id'] == 'amount_due'
        )
        self.assertAlmostEqual(
            next(col['value'] for col in opening['columns']
                 if col['expression_label'] == 'balance'),
            0.001, places=3,
        )
        self.assertAlmostEqual(
            next(col['value'] for col in closing['columns']
                 if col['expression_label'] == 'balance'),
            0.001, places=3,
        )
        self.assertAlmostEqual(total_due, 0.001, places=3)

    def test_open_item_period_settlement_clears_amount_due(self):
        inv = self._invoice(1000.0, '2026-03-10')  # prior period
        # Before the payment the prior invoice is outstanding at date_to.
        self.assertAlmostEqual(
            self._amount_due('2026-06-01', '2026-06-30'), 1000.0, places=2)
        self._pay(inv, '2026-06-15')  # cleared IN the statement period
        # The bug: the clearing payment was excluded, so the running balance
        # kept the 1,000. Fixed: the item drops to residual 0 -> amount due 0,
        # and neither invoice nor settlement survives as a misleading row
        # that merely nets against its counterpart.
        settled = self._statement_payload('2026-06-01', '2026-06-30')
        self.assertAlmostEqual(
            abs(settled['totals']['amount_due']), 0.0, places=2)
        self.assertFalse([
            line for line in settled['lines']
            if str(line.get('id') or '').startswith('aml-')
        ])

    def test_open_item_shows_still_open_invoice(self):
        self._invoice(400.0, '2026-06-05')
        self.assertAlmostEqual(
            self._amount_due('2026-06-01', '2026-06-30'), 400.0, places=2)

    def test_account_type_filter_intersects_customer_and_vendor_scope(self):
        self._invoice(250.0, '2026-06-05')
        self.post_balanced_move(
            [
                {'account': self.account_expense, 'debit': 175.0},
                {
                    'account': self.account_payable,
                    'credit': 175.0,
                    'partner': self.partner,
                    'date_maturity': '2026-06-05',
                },
            ],
            date='2026-06-05',
        )

        self.assertAlmostEqual(
            self._filtered_amount_due(
                self.handler, ['asset_receivable']),
            250.0,
            places=2,
        )
        self.assertAlmostEqual(
            self._filtered_amount_due(
                self.handler, ['liability_payable']),
            0.0,
            places=2,
        )
        self.assertAlmostEqual(
            self._filtered_amount_due(
                self.vendor_handler, ['liability_payable']),
            175.0,
            places=2,
        )
        self.assertAlmostEqual(
            self._filtered_amount_due(
                self.vendor_handler, ['asset_receivable']),
            0.0,
            places=2,
        )

    def test_account_type_filter_applies_to_activity_opening(self):
        self._invoice(125.0, '2025-12-15')
        self.assertAlmostEqual(
            self._filtered_amount_due(
                self.handler, ['asset_receivable'], stype='activity'),
            125.0,
            places=2,
        )
        self.assertAlmostEqual(
            self._filtered_amount_due(
                self.handler, ['liability_payable'], stype='activity'),
            0.0,
            places=2,
        )

    def test_vendor_activity_statement_lists_period_movements(self):
        move = self.post_balanced_move([
            {'account': self.account_expense, 'debit': 175.0},
            {'account': self.account_payable, 'credit': 175.0,
             'partner': self.partner, 'date_maturity': '2026-06-30'},
        ], date='2026-06-05')
        payload = self.vendor_handler.compute({
            'partner_id': self.partner.id,
            'date': {'date_from': '2026-06-01', 'date_to': '2026-06-30'},
            'company_ids': self.env.company.ids,
            'posted_only': True,
            'statement_type': 'activity',
        })
        self.assertEqual(payload['meta']['report_code'], 'vendor_statement')
        self.assertEqual(payload['meta']['statement_type'], 'activity')
        payable_line = move.line_ids.filtered(
            lambda line: line.account_id == self.account_payable)
        self.assertIn(
            'aml-%s' % payable_line.id,
            {line['id'] for line in payload['lines']},
        )
        self.assertAlmostEqual(
            abs(payload['totals']['amount_due']), 175.0, places=2)

    def test_account_type_filtered_xlsx_matches_rendered_rows(self):
        invoice = self._invoice(300.0, '2026-06-05')
        base = {
            'partner_id': self.partner.id,
            'date': {'date_from': '2026-01-01', 'date_to': '2026-12-31'},
            'company_ids': self.env.company.ids,
            'posted_only': True,
            'statement_type': 'open_item',
        }
        matching = self.report.render_xlsx(dict(
            base, account_type_ids=['asset_receivable']))
        excluded = self.report.render_xlsx(dict(
            base, account_type_ids=['liability_payable']))

        def workbook_values(content):
            sheet = load_workbook(io.BytesIO(content)).active
            return {
                cell.value
                for row in sheet.iter_rows()
                for cell in row
                if cell.value not in (None, '')
            }

        self.assertIn(invoice.name, workbook_values(matching))
        self.assertNotIn(invoice.name, workbook_values(excluded))

    def test_partner_print_uses_owner_scoped_transient_attachment(self):
        partner_domain = [
            ('res_model', '=', 'res.partner'),
            ('res_id', '=', self.partner.id),
        ]
        before_partner_attachments = self.env['ir.attachment'].search_count(
            partner_domain)
        with patch.object(
                type(self.report), 'render_pdf', return_value=b'%PDF-1.4\n'):
            action = self.partner.action_print_customer_statement()
        attachment = self.env['ir.attachment'].search([
            ('res_model', '=', 'eh.account.report.wizard'),
            ('name', '=like', 'Customer_Statement_%'),
        ], order='id desc', limit=1)
        self.assertTrue(attachment)
        self.assertIn('/web/content/%s' % attachment.id, action['url'])
        self.assertEqual(attachment.create_uid, self.env.user)
        self.assertEqual(
            self.env['ir.attachment'].search_count(partner_domain),
            before_partner_attachments,
        )

        other = new_test_user(
            self.env,
            login='eh_statement_attachment_other',
            groups='eh_account_base.group_eh_user',
        )
        with self.assertRaises(AccessError):
            attachment.with_user(other).read(['name'])

    def test_partner_vendor_statement_action_uses_vendor_report(self):
        vendor_report = self.env.ref(
            'eh_account_dynamic_reports.report_vendor_statement')
        with patch.object(
                type(vendor_report), 'render_pdf', return_value=b'%PDF-1.4\n'):
            action = self.partner.action_print_vendor_statement()
        attachment = self.env['ir.attachment'].search([
            ('res_model', '=', 'eh.account.report.wizard'),
            ('name', '=like', 'Vendor_Statement_%'),
        ], order='id desc', limit=1)
        self.assertTrue(attachment)
        self.assertIn('/web/content/%s' % attachment.id, action['url'])
