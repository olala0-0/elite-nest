# -*- encoding: utf-8 -*-
##############################################################################
#
# ERP Heritage
# Copyright (C) 2026 (https://www.erpheritage.com.au/)
#
##############################################################################
"""
Deferred revenue / expense schedule handler tests.

The handler reads the optional eh.asset ledger (ships in
eh_account_assets_pro, NOT a dependency of this module). The tests fall
into two groups:

* Unconditional fallback tests: the empty-payload soft-probe path renders
  whether or not the optional module is installed, and the schedule never
  raises.
* Asset-backed tests: skipped when eh.asset is absent. When present, create
  a deferral with a generated schedule spanning the window and assert each
  monthly bucket sums the right recognition-line amounts, the Before bucket
  captures pre-window lines, the Later bucket captures post-window lines,
  Total = sum of buckets = depreciable amount, posted_only behaviour, and
  that the deferred_expense subclass picks only expense-type deferrals.
"""

import io
from datetime import date
from types import SimpleNamespace

from openpyxl import load_workbook

from odoo import fields
from odoo.tests import new_test_user, tagged

from odoo.addons.eh_account_base.tests.common import (
    EhAccountIntegrationTestCase,
)
from odoo.addons.eh_account_base.tools.xlsx_writer import XlsxReportWriter


@tagged('eh_account_dynamic_reports', 'integration', 'post_install',
        '-at_install')
class TestDeferredScheduleHandler(EhAccountIntegrationTestCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.rev_handler = cls.env[
            'eh.account.dynamic.report.handler.deferred_revenue']
        cls.exp_handler = cls.env[
            'eh.account.dynamic.report.handler.deferred_expense']
        cls.report_rev = cls._ensure_report(
            cls, 'deferred_revenue',
            'eh.account.dynamic.report.handler.deferred_revenue')
        cls.asset_available = 'eh.asset' in cls.env

    @staticmethod
    def _ensure_report(cls, code, handler_model):
        report = cls.env['eh.account.dynamic.report'].search(
            [('code', '=', code)], limit=1)
        if not report:
            report = cls.env['eh.account.dynamic.report'].create({
                'code': code, 'name': code,
                'handler_model': handler_model,
            })
        return report

    def setUp(self):
        super().setUp()
        self.options = {
            'date': {'date_from': '2026-01-01', 'date_to': '2026-12-31'},
            'company_ids': [self.company.id],
            'posted_only': False,
        }

    @staticmethod
    def _line_by_id(result, line_id):
        for line in result['lines']:
            if line['id'] == line_id:
                return line
        return None

    @staticmethod
    def _col(line, label):
        if line is None:
            return None
        for col in line['columns']:
            if col['expression_label'] == label:
                return col['value']
        return None

    # ---- unconditional fallback tests ----

    def test_payload_shape_is_well_formed(self):
        result = self.rev_handler.compute(self.options)
        self.assertIn('columns', result)
        self.assertIn('lines', result)
        self.assertIn('totals', result)
        # Before + 12 monthly + Later + Total label columns + the Deferral
        # label column = 1 + 14 + 1 = 16 columns.
        labels = [c['name'] for c in result['columns']]
        self.assertEqual(labels[1], 'Before')
        self.assertEqual(labels[-2], 'Later')
        self.assertEqual(result['columns'][-1]['name'], 'Total')

    def test_module_absent_fallback_is_warning_with_note(self):
        # Patch the soft-probe to report the optional module absent; the
        # handler must degrade to an empty schedule with a 'module not
        # installed' note, never raise. Uses Odoo's patch helper so the
        # override is reverted at test teardown.
        self.patch(
            type(self.rev_handler), '_eh_asset_available',
            lambda self: False)
        result = self.rev_handler.compute(self.options)
        self.assertEqual(len(result['lines']), 1)
        self.assertTrue(result['meta'].get('module_not_installed'))
        self.assertIn('note', result['meta'])
        self.assertEqual(
            result['meta']['warnings'][0]['code'],
            'module_not_installed',
        )
        self.assertEqual(
            result['lines'][0]['name'], result['meta']['note'])
        self.assertEqual(
            result['lines'][0]['meta']['kind'], 'warning')
        self.assertEqual(result['totals']['total'], 0.0)

    def test_provider_failure_is_explicit_without_raw_exception_leak(self):
        self.patch(
            type(self.rev_handler), '_eh_asset_available',
            lambda handler: True,
        )

        def fail_provider(handler, **kwargs):
            raise RuntimeError('secret provider diagnostic')

        self.patch(
            type(self.rev_handler), '_build_schedule_lines', fail_provider)
        result = self.rev_handler.compute(self.options)
        self.assertTrue(result['meta'].get('provider_failed'))
        self.assertEqual(
            result['meta']['warnings'][0]['code'], 'provider_failed')
        self.assertEqual(result['lines'][0]['meta']['kind'], 'warning')
        self.assertNotIn('secret provider diagnostic', str(result))

    def test_provider_warning_is_visible_in_xlsx_and_pdf_rows(self):
        self.patch(
            type(self.rev_handler), '_eh_asset_available',
            lambda handler: False,
        )
        result = self.rev_handler.compute(self.options)
        warning = result['meta']['note']

        workbook = load_workbook(io.BytesIO(
            XlsxReportWriter('Deferred Revenue').write_payload(result)))
        first_column = [
            workbook.active.cell(row=row, column=1).value
            for row in range(1, workbook.active.max_row + 1)
        ]
        self.assertIn(warning, first_column)

        pdf_helper = self.env[
            'report.eh_account_dynamic_reports.report_dynamic_pdf_template'
        ]
        rendered = pdf_helper._render_lines(result)
        self.assertIn(warning, [line['name'] for line in rendered])

    def test_bucket_resolution_monthly(self):
        buckets = self.rev_handler._resolve_period_buckets(
            date(2026, 1, 1), date(2026, 12, 31))
        self.assertEqual(len(buckets), 12)
        self.assertEqual(buckets[0]['label'], '2026-01')
        self.assertEqual(buckets[-1]['label'], '2026-12')

    def test_asset_discovery_domain_excludes_draft_and_cancelled_states(self):
        fake_asset_model = SimpleNamespace(_fields={
            'state': SimpleNamespace(selection=[
                ('draft', 'Draft'), ('active', 'Running'),
                ('disposed', 'Disposed'), ('cancelled', 'Cancelled'),
            ]),
        })
        domain = self.rev_handler._asset_discovery_domain(
            fake_asset_model, [self.company.id])
        self.assertEqual(domain, [
            ('deferred_type', '=', 'deferred_revenue'),
            ('company_id', 'in', [self.company.id]),
            ('state', 'not in', ['cancelled', 'draft']),
        ])

    def test_bucket_resolution_coarsens_for_wide_window(self):
        # 5-year window -> quarterly buckets (more than 36 months).
        buckets = self.rev_handler._resolve_period_buckets(
            date(2026, 1, 1), date(2030, 12, 31))
        self.assertTrue(all('Q' in b['label'] for b in buckets))
        self.assertLessEqual(len(buckets), 36)

    def test_bucket_index_edges(self):
        buckets = self.rev_handler._resolve_period_buckets(
            date(2026, 1, 1), date(2026, 12, 31))
        df, dt = date(2026, 1, 1), date(2026, 12, 31)
        # Before window.
        self.assertEqual(
            self.rev_handler._bucket_index_for_date(
                date(2025, 12, 1), df, dt, buckets), 0)
        # After window.
        self.assertEqual(
            self.rev_handler._bucket_index_for_date(
                date(2027, 3, 1), df, dt, buckets), len(buckets) + 1)
        # March -> third in-window bucket (index 3).
        self.assertEqual(
            self.rev_handler._bucket_index_for_date(
                date(2026, 3, 15), df, dt, buckets), 3)

    def test_posted_only_excludes_unposted_schedule_lines(self):
        fake_asset = SimpleNamespace(
            id=987001,
            display_name='Synthetic Deferred Revenue',
            depreciation_account_id=SimpleNamespace(id=456001),
            depreciation_line_ids=[
                SimpleNamespace(
                    depreciation_date=date(2026, 1, 31),
                    amount=100.0,
                    is_posted=True,
                ),
                SimpleNamespace(
                    depreciation_date=date(2026, 2, 28),
                    amount=200.0,
                    is_posted=False,
                ),
            ],
        )
        self.patch(
            type(self.rev_handler), '_eh_asset_available',
            lambda handler: True,
        )
        self.patch(
            type(self.rev_handler), '_discover_assets',
            lambda handler, company_ids: [fake_asset],
        )

        all_schedule = self.rev_handler.compute(dict(
            self.options, posted_only=False))
        posted_schedule = self.rev_handler.compute(dict(
            self.options, posted_only=True))
        all_row = self._line_by_id(all_schedule, 'asset-987001')
        posted_row = self._line_by_id(posted_schedule, 'asset-987001')
        self.assertAlmostEqual(self._col(all_row, 'total'), 300.0, places=2)
        self.assertAlmostEqual(
            self._col(posted_row, 'total'), 100.0, places=2)
        self.assertTrue(posted_row['meta']['has_unposted'])

    def test_posted_only_excludes_wholly_unposted_deferral(self):
        fake_asset = SimpleNamespace(
            id=987002,
            display_name='Unposted Deferred Revenue',
            depreciation_account_id=SimpleNamespace(id=456002),
            depreciation_line_ids=[SimpleNamespace(
                depreciation_date=date(2026, 1, 31),
                amount=100.0,
                is_posted=False,
            )],
        )
        self.patch(
            type(self.rev_handler), '_eh_asset_available',
            lambda handler: True,
        )
        self.patch(
            type(self.rev_handler), '_discover_assets',
            lambda handler, company_ids: [fake_asset],
        )

        result = self.rev_handler.compute(dict(
            self.options, posted_only=True))
        self.assertIsNone(self._line_by_id(result, 'asset-987002'))
        self.assertAlmostEqual(result['totals']['total'], 0.0, places=2)

    def test_kwd_precision_keeps_one_fils_schedule_bucket(self):
        kwd = self.env['res.currency'].with_context(active_test=False).search(
            [('name', '=', 'KWD')], limit=1,
        )
        self.assertTrue(kwd)
        fake_asset = SimpleNamespace(
            id=987003,
            display_name='KWD Deferred Revenue',
            depreciation_account_id=SimpleNamespace(id=456003),
            depreciation_line_ids=[SimpleNamespace(
                depreciation_date=date(2026, 1, 31),
                amount=0.001,
                is_posted=True,
            )],
        )
        self.patch(
            type(self.rev_handler), '_discover_assets',
            lambda handler, company_ids: [fake_asset],
        )
        date_from = date(2026, 1, 1)
        date_to = date(2026, 12, 31)
        buckets = self.rev_handler._resolve_period_buckets(
            date_from, date_to,
        )
        lines, total = self.rev_handler._build_schedule_lines(
            options=self.options, company_ids=[self.company.id],
            date_from=date_from, date_to=date_to, posted_only=True,
            buckets=buckets, currency=kwd,
        )
        self.assertEqual(len(lines), 1)
        self.assertAlmostEqual(self._col(lines[0], 'period_2'), 0.001, places=3)
        self.assertAlmostEqual(self._col(lines[0], 'total'), 0.001, places=3)
        self.assertAlmostEqual(total, 0.001, places=3)

    # ---- asset-backed tests ----

    def _make_deferred(self, deferred_type):
        Category = self.env['eh.asset.category']
        holding_type = ('liability_current'
                        if deferred_type == 'deferred_revenue'
                        else 'asset_current')
        holding = self._ensure_account(
            self.env, '2400' if deferred_type == 'deferred_revenue'
            else '1410',
            'Deferred Holding', holding_type)
        recognition = (self.account_revenue
                       if deferred_type == 'deferred_revenue'
                       else self.account_expense)
        accum = self._ensure_account(
            self.env, '1599', 'Accum', 'asset_fixed')
        category = Category.create({
            'name': 'Deferral cat %s' % deferred_type,
            'method': 'straight_line',
            'useful_life_months': 12,
            'asset_account_id': holding.id,
            'depreciation_account_id': recognition.id,
            'accumulated_depreciation_account_id': accum.id,
            'journal_id': self.journal_misc.id,
            'company_id': self.company.id,
        })
        asset = self.env['eh.asset'].create({
            'name': '/',
            'category_id': category.id,
            'deferred_type': deferred_type,
            'partner_id': self.partner_a.id,
            'acquisition_date': '2026-01-01',
            'in_service_date': '2026-01-31',
            'acquisition_cost': 12000.0,
            'salvage_value': 0.0,
            'method': 'straight_line',
            'useful_life_months': 12,
            'prorate_first_period': False,
            'asset_account_id': holding.id,
            'depreciation_account_id': recognition.id,
            'accumulated_depreciation_account_id': accum.id,
            'journal_id': self.journal_misc.id,
        })
        asset.action_compute_schedule()
        asset.action_activate()
        return asset

    def test_schedule_buckets_and_total(self):
        if not self.asset_available:
            self.skipTest("eh_account_assets_pro not installed")
        asset = self._make_deferred('deferred_revenue')
        result = self.rev_handler.compute(self.options)
        row = self._line_by_id(result, 'asset-%s' % asset.id)
        self.assertIsNotNone(row)

        # Total = depreciable amount = acquisition cost = 12000.
        total = self._col(row, 'total')
        self.assertAlmostEqual(total, 12000.0, places=2)

        # Sum of all period buckets equals the row total.
        bucket_sum = sum(
            self._col(row, 'period_%d' % i) or 0.0
            for i in range(1, len(result['columns']) - 1))
        self.assertAlmostEqual(bucket_sum, total, places=2)

        # The grand total foots to the same figure.
        self.assertAlmostEqual(result['totals']['total'], 12000.0, places=2)

    def test_before_and_later_buckets(self):
        if not self.asset_available:
            self.skipTest("eh_account_assets_pro not installed")
        asset = self._make_deferred('deferred_revenue')
        # Narrow the window to Feb-Mar 2026 so Jan lands in Before and
        # Apr-Dec land in Later.
        opts = dict(self.options, date={
            'date_from': '2026-02-01', 'date_to': '2026-03-31'})
        result = self.rev_handler.compute(opts)
        row = self._line_by_id(result, 'asset-%s' % asset.id)
        self.assertIsNotNone(row)
        before = self._col(row, 'period_1')  # Before
        later = self._col(row, 'period_%d' % (len(result['columns']) - 2))
        # One 1,000 monthly recognition is before the window; nine are after.
        self.assertAlmostEqual(before, 1000.0, places=2)
        self.assertAlmostEqual(later, 9000.0, places=2)
        # Total still foots to the full 12000.
        self.assertAlmostEqual(self._col(row, 'total'), 12000.0, places=2)

    def test_drilldown_matches_before_period_later_and_total_cells(self):
        if not self.asset_available:
            self.skipTest("eh_account_assets_pro not installed")
        asset = self._make_deferred('deferred_revenue')
        manager = new_test_user(
            self.env,
            login='eh_deferred_drilldown_manager',
            groups='eh_account_base.group_eh_manager',
        )
        for schedule_line in asset.depreciation_line_ids:
            schedule_line.with_user(manager).action_post()
        opts = dict(self.options, date={
            'date_from': '2026-02-01', 'date_to': '2026-03-31',
        })
        line_id = 'asset-%s' % asset.id

        def dates_for(expression_label):
            action = self.rev_handler.get_drilldown_action(
                dict(opts, _eh_column_expression=expression_label),
                line_id,
            )
            return set(self.env['account.move.line'].search(
                action['domain'],
            ).mapped('date'))

        before_dates = dates_for('period_1')
        february_dates = dates_for('period_2')
        later_dates = dates_for('period_4')
        total_dates = dates_for('total')
        self.assertTrue(before_dates)
        self.assertTrue(all(value < date(2026, 2, 1)
                            for value in before_dates))
        self.assertTrue(february_dates)
        self.assertTrue(all(date(2026, 2, 1) <= value <= date(2026, 2, 28)
                            for value in february_dates))
        self.assertTrue(later_dates)
        self.assertTrue(all(value > date(2026, 3, 31)
                            for value in later_dates))
        self.assertTrue(before_dates | february_dates | later_dates
                        <= total_dates)

    def test_deferred_expense_subclass_scope(self):
        if not self.asset_available:
            self.skipTest("eh_account_assets_pro not installed")
        rev_asset = self._make_deferred('deferred_revenue')
        exp_asset = self._make_deferred('deferred_expense')
        exp_result = self.exp_handler.compute(self.options)
        ids = {l['id'] for l in exp_result['lines']}
        self.assertIn('asset-%s' % exp_asset.id, ids)
        self.assertNotIn('asset-%s' % rev_asset.id, ids)

    def test_empty_when_no_deferrals(self):
        if not self.asset_available:
            self.skipTest("eh_account_assets_pro not installed")
        # No deferrals created in this test -> empty line list, no raise.
        result = self.rev_handler.compute(self.options)
        self.assertEqual(
            [l for l in result['lines'] if l['id'].startswith('asset-')], [])

    def test_restricted_user_without_asset_acl_gets_warning_not_sudo_data(self):
        if not self.asset_available:
            self.skipTest("eh_account_assets_pro not installed")
        asset = self._make_deferred('deferred_revenue')
        restricted_user = new_test_user(
            self.env,
            login='eh_deferred_provider_restricted_user',
            groups='base.group_user',
        )

        result = self.rev_handler.with_user(
            restricted_user,
        ).compute(self.options)
        self.assertTrue(result['meta'].get('provider_failed'))
        self.assertFalse(any(
            line.get('id') == 'asset-%s' % asset.id
            for line in result['lines']
        ))
