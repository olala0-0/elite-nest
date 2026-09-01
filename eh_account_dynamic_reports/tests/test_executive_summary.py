# -*- encoding: utf-8 -*-
##############################################################################
#
# ERP Heritage
# Copyright (C) 2026 (https://www.erpheritage.com.au/)
#
##############################################################################
"""
Executive Summary handler tests.

Seeds income / expense / cash / AR / AP and asserts:

* Revenue and Net Profit reconcile to the dashboard's period_revenue /
  period_net for the same window (regression lock against the board).
* Margins = profit / revenue; current ratio = current assets / current
  liabilities; DSO math.
* Divide-by-zero ratios return a safe sentinel (n/a), never raise.
* The comparison column is populated when a comparison is set.
* Orchestrator wiring: the report record renders without error.
"""

import io

from openpyxl import load_workbook

from odoo import fields
from odoo.exceptions import UserError
from odoo.tests import tagged

from odoo.addons.eh_account_base.tests.common import (
    EhAccountIntegrationTestCase,
)


@tagged('eh_account_dynamic_reports', 'integration', 'post_install',
        '-at_install')
class TestExecutiveSummaryHandler(EhAccountIntegrationTestCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.handler = cls.env[
            'eh.account.dynamic.report.handler.executive_summary'
        ]
        cls.account_direct_cost = cls._ensure_account(
            cls.env, '5100', 'Direct Costs', 'expense_direct_cost')
        cls.report = cls.env['eh.account.dynamic.report'].search(
            [('code', '=', 'executive_summary')], limit=1)
        if not cls.report:
            cls.report = cls.env['eh.account.dynamic.report'].create({
                'code': 'executive_summary',
                'name': 'Executive Summary',
                'handler_model':
                    'eh.account.dynamic.report.handler.executive_summary',
            })

    def setUp(self):
        super().setUp()
        self.options = {
            'date': {'date_from': '2026-01-01', 'date_to': '2026-12-31'},
            'company_ids': [self.company.id],
            'posted_only': True,
        }

    def _post(self, lines, date_str='2026-06-15'):
        return self.post_balanced_move(
            lines, date=fields.Date.from_string(date_str))

    def test_column_axis_capability_contract_is_comparison_only(self):
        self.assertEqual(
            self.handler._EH_COLUMN_AXIS_CAPABILITIES,
            frozenset({'comparison'}),
        )
        normalized = self.handler.normalize_options(dict(
            self.options,
            comparison='previous_year',
            comparison_number=2,
            comparison_order='ascending',
        ))
        self.assertEqual(normalized['comparison_number'], 2)
        self.assertNotIn('analytic_column_account_ids', normalized)
        with self.assertRaisesRegex(
                UserError, 'does not support.*analytic columns'):
            self.handler.normalize_options(dict(
                self.options, analytic_column_account_ids=[1],
            ))

    @staticmethod
    def _line_by_id(result, line_id):
        for line in result['lines']:
            if line['id'] == line_id:
                return line
        return None

    @staticmethod
    def _value(line):
        if line is None:
            return None
        for col in line['columns']:
            if col['expression_label'] == 'value':
                return col['value']
        return None

    @staticmethod
    def _cell(line, expression_label='value'):
        if line:
            for col in line['columns']:
                if col['expression_label'] == expression_label:
                    return col
        return {}

    def _seed_full_set(self):
        # Revenue 10,000 (Dr AR / Cr Revenue).
        self._post([
            {'account': self.account_receivable, 'debit': 10000.0,
             'partner': self.partner_a},
            {'account': self.account_revenue, 'credit': 10000.0},
        ])
        # Expense 4,000 (Dr Expense / Cr AP).
        self._post([
            {'account': self.account_expense, 'debit': 4000.0},
            {'account': self.account_payable, 'credit': 4000.0,
             'partner': self.partner_b},
        ])
        # Cash receipt 6,000 (Dr Cash / Cr AR), partial collection.
        self._post([
            {'account': self.account_cash, 'debit': 6000.0},
            {'account': self.account_receivable, 'credit': 6000.0,
             'partner': self.partner_a},
        ])

    def test_depreciation_reduces_operating_profit(self):
        depreciation = self.env['account.account'].create({
            'code': '6105',
            'name': 'Depreciation Expense',
            'account_type': 'expense_depreciation',
            'company_ids': [(6, 0, [self.company.id])],
        })
        self._post([
            {'account': self.account_cash, 'debit': 1000.0},
            {'account': self.account_revenue, 'credit': 1000.0},
        ])
        self._post([
            {'account': self.account_expense, 'debit': 100.0},
            {'account': self.account_cash, 'credit': 100.0},
        ])
        self._post([
            {'account': depreciation, 'debit': 50.0},
            {'account': self.account_cash, 'credit': 50.0},
        ])
        result = self.handler.compute(self.options)
        self.assertAlmostEqual(
            self._value(self._line_by_id(result, 'exec-operating_profit')),
            850.0, places=2)

    def test_cost_of_sales_and_total_expense_are_report_rows(self):
        self._post([
            {'account': self.account_cash, 'debit': 1000.0},
            {'account': self.account_revenue, 'credit': 1000.0},
        ])
        self._post([
            {'account': self.account_direct_cost, 'debit': 100.0},
            {'account': self.account_expense, 'debit': 50.0},
            {'account': self.account_cash, 'credit': 150.0},
        ])

        result = self.handler.compute(self.options)
        cost = self._line_by_id(result, 'exec-cost_of_sales')
        total = self._line_by_id(result, 'exec-total_expense')
        self.assertAlmostEqual(self._value(cost), 100.0, places=2)
        self.assertAlmostEqual(self._value(total), 150.0, places=2)
        self.assertEqual(self._cell(cost)['figure_type'], 'monetary')
        self.assertEqual(self._cell(total)['figure_type'], 'monetary')
        self.assertEqual(cost['meta']['metric'], 'cost_of_sales')
        self.assertEqual(total['meta']['metric'], 'total_expense')

    # ---- reconciliation regression lock vs dashboard ----

    def test_kwd_precision_preserves_lines_totals_and_half_step_tie(self):
        kwd = self.env.ref('base.KWD')
        self.assertEqual(kwd.decimal_places, 3)

        amounts = {
            self.handler.INCOME_TYPES: 1.2345,
            self.handler.EXPENSE_TYPES: 1.2335,
            self.handler.CURRENT_ASSET_TYPES: 1.2345,
            self.handler.CURRENT_LIABILITY_TYPES: 1.2335,
        }

        def fake_balance(handler, account_types=None, **kwargs):
            return amounts.get(tuple(account_types or ()), 0.0)

        self.patch(
            type(self.handler), '_fetch_aggregate_balance', fake_balance)
        scalars = self.handler._compute_scalars(
            options=self.options,
            company_ids=[self.company.id],
            date_from=fields.Date.to_date('2026-01-01'),
            date_to=fields.Date.to_date('2026-12-31'),
            posted_only=True,
            currency=kwd,
        )
        lines = self.handler._build_lines(scalars, prior=None)
        payload = {'lines': lines}

        revenue = self._line_by_id(payload, 'exec-revenue')
        net_profit = self._line_by_id(payload, 'exec-net_profit')
        self.assertEqual(self._value(revenue), 1.235)
        self.assertEqual(self._value(net_profit), 0.001)
        self.assertEqual(scalars['revenue'], 1.235)
        self.assertEqual(scalars['net_profit'], 0.001)
        self.assertAlmostEqual(
            scalars['net_margin'], 0.001 / 1.2345, places=9)

    def test_revenue_and_net_reconcile_to_dashboard(self):
        self._seed_full_set()
        result = self.handler.compute(self.options)

        revenue = self._value(self._line_by_id(result, 'exec-revenue'))
        net = self._value(self._line_by_id(result, 'exec-net_profit'))
        self.assertAlmostEqual(revenue, 10000.0, places=2)
        self.assertAlmostEqual(net, 6000.0, places=2)

        # Cross-check against the live dashboard for the same window when
        # the dashboard module is installed (it is an optional sibling, not
        # a dependency of this module).
        if 'eh.account.dashboard' in self.env:
            board = self.env['eh.account.dashboard'].create({
                'company_id': self.company.id,
                'posted_only': True,
                'period_date_from': '2026-01-01',
                'period_date_to': '2026-12-31',
            })
            self.assertAlmostEqual(revenue, board.period_revenue, places=2)
            self.assertAlmostEqual(net, board.period_net, places=2)

    def test_margins_and_balances(self):
        self._seed_full_set()
        result = self.handler.compute(self.options)

        # Net margin = net / revenue = 6000 / 10000 = 0.6.
        net_margin = self._value(self._line_by_id(result, 'exec-net_margin'))
        self.assertAlmostEqual(net_margin, 0.6, places=4)

        # Cash balance cumulative to date_to = 6000.
        cash = self._value(self._line_by_id(result, 'exec-cash'))
        self.assertAlmostEqual(cash, 6000.0, places=2)
        # Receivables remaining = 10000 - 6000 = 4000.
        receivables = self._value(
            self._line_by_id(result, 'exec-receivables'))
        self.assertAlmostEqual(receivables, 4000.0, places=2)
        # Payables = 4000 (presented positive).
        payables = self._value(self._line_by_id(result, 'exec-payables'))
        self.assertAlmostEqual(payables, 4000.0, places=2)

    def test_current_ratio_math(self):
        self._seed_full_set()
        result = self.handler.compute(self.options)
        # Current assets = cash 6000 + AR 4000 = 10000.
        # Current liabilities = AP 4000.
        # Current ratio = 10000 / 4000 = 2.5.
        current_ratio = self._value(
            self._line_by_id(result, 'exec-current_ratio'))
        self.assertAlmostEqual(current_ratio, 2.5, places=4)

    def test_dso_math(self):
        self._seed_full_set()
        result = self.handler.compute(self.options)
        # DSO = receivables / revenue * period_days
        #     = 4000 / 10000 * 365 = 146.0.
        dso = self._value(self._line_by_id(result, 'exec-dso'))
        self.assertAlmostEqual(dso, 146.0, places=0)

    # ---- divide-by-zero safety ----

    def test_zero_revenue_ratios_are_safe(self):
        # No data at all: every ratio is undefined, none raises.
        result = self.handler.compute(self.options)
        for line_id in (
            'exec-net_margin', 'exec-gross_margin', 'exec-operating_margin',
            'exec-current_ratio', 'exec-quick_ratio', 'exec-dso', 'exec-dpo',
            'exec-return_on_assets',
        ):
            line = self._line_by_id(result, line_id)
            self.assertIsNotNone(line)
            # Numeric semantics remain declared, but undefined is not zero.
            self.assertEqual(self._value(line), 'n/a')
            self.assertIn(
                self._cell(line).get('figure_type'),
                ('percentage', 'float', 'integer'),
            )

    def test_safe_ratio_helper(self):
        self.assertIsNone(self.handler._safe_ratio(5.0, 0))
        self.assertIsNone(self.handler._safe_ratio(5.0, None))
        self.assertAlmostEqual(
            self.handler._safe_ratio(6.0, 3.0), 2.0, places=4)

    # ---- comparison column ----

    def test_comparison_column_populated(self):
        # Prior year revenue 8000.
        self._post([
            {'account': self.account_receivable, 'debit': 8000.0,
             'partner': self.partner_a},
            {'account': self.account_revenue, 'credit': 8000.0},
        ], date_str='2025-06-15')
        self._seed_full_set()
        result = self.handler.compute(dict(
            self.options, comparison='previous_year'))
        # Three columns now: metric + value + prior_value.
        labels = [c['expression_label'] for c in result['columns']]
        self.assertIn('prior_value', labels)
        revenue_line = self._line_by_id(result, 'exec-revenue')
        prior_col = [c for c in revenue_line['columns']
                     if c['expression_label'] == 'prior_value']
        self.assertEqual(len(prior_col), 1)
        self.assertAlmostEqual(prior_col[0]['value'], 8000.0, places=2)
        self.assertEqual(prior_col[0]['figure_type'], 'monetary')
        total_expense = self._line_by_id(result, 'exec-total_expense')
        self.assertAlmostEqual(self._value(total_expense), 4000.0, places=2)
        prior_expense = self._cell(total_expense, 'prior_value')
        self.assertAlmostEqual(prior_expense['value'], 0.0, places=2)
        self.assertEqual(prior_expense['figure_type'], 'monetary')

    def test_n_period_axis_orders_complete_kpi_columns_ascending(self):
        def fake_scalars(handler, date_to=None, **_kwargs):
            value = float(date_to.year)
            return {'revenue': value, 'net_profit': value}

        self.patch(type(self.handler), '_compute_scalars', fake_scalars)
        result = self.handler.compute(dict(
            self.options,
            comparison='previous_year',
            comparison_number=2,
            comparison_order='ascending',
        ))
        value_columns = result['columns'][1:]
        self.assertEqual(
            [column['scope']['date_to'] for column in value_columns],
            ['2024-12-31', '2025-12-31', '2026-12-31'],
        )
        self.assertTrue(all(column.get('scope') for column in value_columns))
        revenue = self._line_by_id(result, 'exec-revenue')
        self.assertEqual(
            [column['value'] for column in revenue['columns']],
            [2024.0, 2025.0, 2026.0],
        )
        gross_margin = self._line_by_id(result, 'exec-gross_margin')
        self.assertEqual(
            [column['value'] for column in gross_margin['columns']],
            ['n/a', 'n/a', 'n/a'],
        )

    def test_custom_comparison_scope_uses_requested_window(self):
        def fake_scalars(handler, date_from=None, date_to=None, **_kwargs):
            value = float((date_to - date_from).days + 1)
            return {'revenue': value, 'net_profit': value}

        self.patch(type(self.handler), '_compute_scalars', fake_scalars)
        result = self.handler.compute(dict(
            self.options,
            comparison='custom',
            comparison_custom_date_from='2024-02-01',
            comparison_custom_date_to='2024-02-29',
        ))
        value_columns = result['columns'][1:]
        self.assertEqual(
            [(column['scope']['date_from'], column['scope']['date_to'])
             for column in value_columns],
            [
                ('2026-01-01', '2026-12-31'),
                ('2024-02-01', '2024-02-29'),
            ],
        )
        revenue = self._line_by_id(result, 'exec-revenue')
        self.assertEqual(
            [column['value'] for column in revenue['columns']],
            [365.0, 29.0],
        )

    def test_each_value_cell_declares_numeric_semantics(self):
        self._seed_full_set()
        result = self.handler.compute(self.options)
        expected = {
            'exec-revenue': 'monetary',
            'exec-gross_margin': 'percentage',
            'exec-current_ratio': 'float',
            'exec-dso': 'integer',
        }
        for line_id, figure_type in expected.items():
            cell = self._cell(self._line_by_id(result, line_id))
            self.assertEqual(cell['figure_type'], figure_type)
            self.assertIsInstance(cell['value'], (int, float))

    def test_presentation_currency_converts_only_monetary_cells_and_xlsx(self):
        target = self.env['res.currency'].create({
            'name': 'ZXE', 'symbol': 'X', 'rounding': 0.01,
        })
        self.env['res.currency.rate'].create({
            'currency_id': target.id,
            'name': '2026-01-01',
            'rate': 2.0,
            'company_id': self.company.id,
        })
        self._seed_full_set()
        options = dict(
            self.options, presentation_currency_id=target.id,
        )

        payload = self.report.render(options, use_cache=False)
        revenue_cell = self._cell(
            self._line_by_id(payload, 'exec-revenue'))
        margin_cell = self._cell(
            self._line_by_id(payload, 'exec-net_margin'))
        self.assertEqual(revenue_cell['figure_type'], 'monetary')
        self.assertAlmostEqual(revenue_cell['value'], 20000.0, places=2)
        self.assertEqual(margin_cell['figure_type'], 'percentage')
        self.assertAlmostEqual(margin_cell['value'], 0.6, places=4)
        self.assertEqual(payload['currency']['id'], target.id)

        content = self.report.render_xlsx(options, use_cache=False)
        sheet = load_workbook(io.BytesIO(content), data_only=True).active
        revenue_row = next(
            row for row in range(1, sheet.max_row + 1)
            if sheet.cell(row=row, column=1).value == 'Revenue'
        )
        margin_row = next(
            row for row in range(1, sheet.max_row + 1)
            if sheet.cell(row=row, column=1).value == 'Net Margin'
        )
        revenue_xlsx = sheet.cell(row=revenue_row, column=2)
        margin_xlsx = sheet.cell(row=margin_row, column=2)
        self.assertAlmostEqual(revenue_xlsx.value, 20000.0, places=2)
        self.assertIn('X', revenue_xlsx.number_format)
        self.assertEqual(revenue_xlsx.alignment.horizontal, 'right')
        self.assertAlmostEqual(margin_xlsx.value, 0.6, places=4)
        self.assertEqual(margin_xlsx.number_format, '0.00%')

    def test_axis_currency_uses_each_period_flow_average_and_closing_rate(self):
        target = self.env['res.currency'].create({
            'name': 'ZXA', 'symbol': 'A', 'rounding': 0.01,
        })
        self.env['res.currency.rate'].create([
            {
                'currency_id': target.id,
                'name': '2025-01-01',
                'rate': 2.0,
                'company_id': self.company.id,
            },
            {
                'currency_id': target.id,
                'name': '2026-01-01',
                'rate': 4.0,
                'company_id': self.company.id,
            },
        ])
        for date_str in ('2025-06-15', '2026-06-15'):
            self._post([
                {
                    'account': self.account_receivable,
                    'debit': 100.0,
                    'partner': self.partner_a,
                },
                {'account': self.account_revenue, 'credit': 100.0},
            ], date_str=date_str)

        payload = self.report.render(dict(
            self.options,
            presentation_currency_id=target.id,
            comparison='custom',
            comparison_custom_date_from='2025-01-01',
            comparison_custom_date_to='2025-12-31',
        ), use_cache=False)
        revenue = self._line_by_id(payload, 'exec-revenue')
        receivables = self._line_by_id(payload, 'exec-receivables')
        self.assertEqual(
            [cell['value'] for cell in revenue['columns']],
            [400.0, 200.0],
        )
        # Snapshot cells include cumulative receivables at each scope end:
        # current 200 * 4.0; prior 100 * 2.0.
        self.assertEqual(
            [cell['value'] for cell in receivables['columns']],
            [800.0, 200.0],
        )
        self.assertEqual(payload['currency']['id'], target.id)
        self.assertEqual(
            payload['meta']['currency_translation_policy'],
            'mixed_flow_average_and_closing_spot',
        )
        periods = payload['meta']['currency_translation_periods']
        self.assertEqual([period['label'] for period in periods], [
            'amount__period_current',
            'amount__period_comparison_1',
        ])
        self.assertEqual(
            [period['flow']['policy'] for period in periods],
            ['period_average', 'period_average'],
        )
        self.assertEqual(
            [period['snapshot']['policy'] for period in periods],
            ['closing_spot', 'closing_spot'],
        )

    # ---- orchestrator wiring ----

    def test_report_renders_through_orchestrator(self):
        self._seed_full_set()
        payload = self.report.render(self.options)
        self.assertIn('lines', payload)
        self.assertIn('columns', payload)
        self.assertTrue(any(
            l['id'] == 'exec-revenue' for l in payload['lines']))

    # ---- drilldown ----

    def test_cash_row_drills_to_journal_items(self):
        self._seed_full_set()
        action = self.handler.get_drilldown_action(self.options, 'exec-cash')
        self.assertIsInstance(action, dict)
        self.assertEqual(action['res_model'], 'account.move.line')
        # Ratio rows do not drill.
        self.assertIsNone(
            self.handler.get_drilldown_action(
                self.options, 'exec-current_ratio'))
