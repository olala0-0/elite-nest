# -*- encoding: utf-8 -*-
##############################################################################
#
# ERP Heritage
# Copyright (C) 2026 (https://www.erpheritage.com.au/)
#
##############################################################################
"""
General Ledger handler tests.

Covers:

* Account header line, opening line, and total line render correctly.
* Entries within the period appear with date, journal, move, partner,
  label, debit, credit, and running balance.
* Running balance: opening + sum(debit) - sum(credit) per account.
* Opening balance reflects entries strictly before date_from.
* Closing balance equals running balance after the last entry.
* Out of period entries excluded.
* Posted only excludes draft entries; setting it false includes them.
* Cancelled entries excluded.
* Account, journal, partner filters narrow the result set.
* Missing dates raise UserError.
* Orchestrator render works and respects the cache.
* Drill down: aml-X opens the specific journal entry form.
* Drill down: account-N opens journal items filtered to that account.
* Drill down: opening / total / unrelated ids return None.
* XLSX export produces a valid workbook.
"""

from io import BytesIO
from unittest.mock import patch

from openpyxl import load_workbook

from odoo import fields
from odoo.exceptions import AccessError, UserError
from odoo.tests import new_test_user, tagged
from odoo.tools import SQL

from odoo.addons.eh_account_base.tests.common import EhAccountIntegrationTestCase


@tagged('eh_account_dynamic_reports', 'integration', 'post_install', '-at_install')
class TestGeneralLedgerHandler(EhAccountIntegrationTestCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.handler = cls.env[
            'eh.account.dynamic.report.handler.general_ledger'
        ]
        cls.report = cls.env['eh.account.dynamic.report'].search(
            [('code', '=', 'general_ledger')], limit=1,
        )
        if not cls.report:
            cls.report = cls.env['eh.account.dynamic.report'].create({
                'code': 'general_ledger',
                'name': 'General Ledger',
                'handler_model':
                    'eh.account.dynamic.report.handler.general_ledger',
            })

    def setUp(self):
        super().setUp()
        self.options = {
            'date': {'date_from': '2026-01-01', 'date_to': '2026-12-31'},
            'company_ids': [self.company.id],
            'posted_only': True,
            'show_zero': False,
        }

    def _post_in_period(self, lines, date_str='2026-06-15'):
        return self.post_balanced_move(
            lines, date=fields.Date.from_string(date_str),
        )

    def test_column_axis_capability_contract_rejects_direct_rpc_options(self):
        self.assertFalse(self.handler._EH_COLUMN_AXIS_CAPABILITIES)
        with self.assertRaisesRegex(
                UserError, 'does not support.*comparison columns'):
            self.handler.normalize_options(dict(
                self.options, comparison='previous_year',
            ))
        with self.assertRaisesRegex(
                UserError, 'does not support.*analytic columns'):
            self.handler.normalize_options(dict(
                self.options, analytic_column_account_ids=[1],
            ))

    @staticmethod
    def _lines_for_account(result, account_id):
        return [
            line for line in result['lines']
            if (line.get('meta') or {}).get('account_id') == account_id
        ]

    @staticmethod
    def _column_value(line, label):
        for col in line['columns']:
            if col['expression_label'] == label:
                return col['value']
        return None

    @staticmethod
    def _line_by_kind(lines, kind):
        for line in lines:
            if (line.get('meta') or {}).get('kind') == kind:
                return line
        return None

    def test_foreign_currency_amount_on_line(self):
        currency = self.env['res.currency'].create({
            'name': 'ZZG', 'symbol': 'G',
            'rounding': self.company.currency_id.rounding,
        })
        self.env['res.currency.rate'].create({
            'currency_id': currency.id, 'name': '2026-01-01',
            'rate': 0.5, 'company_id': self.company.id})
        move = self.env['account.move'].create({
            'move_type': 'entry',
            'date': fields.Date.from_string('2026-06-15'),
            'journal_id': self.journal_misc.id,
            'line_ids': [
                (0, 0, {'account_id': self.account_revenue.id,
                        'debit': 0.0, 'credit': 110.0,
                        'currency_id': currency.id,
                        'amount_currency': -200.0}),
                (0, 0, {'account_id': self.account_cash.id,
                        'debit': 110.0, 'credit': 0.0}),
            ],
        })
        move.action_post()

        result = self.handler.compute(self.options)
        rev_aml = self._line_by_kind(
            self._lines_for_account(result, self.account_revenue.id), 'aml')
        self.assertEqual(
            self._column_value(rev_aml, 'foreign'), '-200.00 ZZG')
        # A company-currency line shows no foreign amount.
        cash_aml = self._line_by_kind(
            self._lines_for_account(result, self.account_cash.id), 'aml')
        self.assertEqual(self._column_value(cash_aml, 'foreign'), '')

    def test_eager_export_uses_hard_ceiling_not_company_soft_cap(self):
        original = self.company.eh_gl_row_limit
        original_page_size = self.company.eh_expand_page_size
        self.addCleanup(self.company.write, {
            'eh_gl_row_limit': original,
            'eh_expand_page_size': original_page_size,
        })
        self.company.eh_gl_row_limit = 10_000
        self.assertEqual(self.handler._resolve_row_limit({}), 10_000)
        self.assertEqual(
            self.handler._resolve_row_limit({'eager_expand': True}),
            100_000,
        )
        self.assertEqual(
            self.handler._resolve_row_limit({
                'eager_expand': True, 'row_limit': 25_000,
            }),
            25_000,
        )
        self.company.write({
            'eh_gl_row_limit': 1,
            'eh_expand_page_size': 1,
        })
        self._post_in_period([
            {'account': self.account_cash, 'debit': 25.0},
            {'account': self.account_revenue, 'credit': 25.0},
        ])
        with self.assertRaises(UserError):
            self.handler.compute(self.options)
        exported = self.handler.compute(dict(
            self.options, eager_expand=True,
        ))
        self.assertGreaterEqual(
            len([line for line in exported['lines']
                 if (line.get('meta') or {}).get('kind') == 'aml']),
            2,
        )

    def test_render_clamps_hidden_company_before_ledger_query(self):
        hidden_company = self.env['res.company'].create({
            'name': 'Hidden GL Company',
        })
        hidden_env = self.env['account.move'].with_context(
            allowed_company_ids=[hidden_company.id],
        ).with_company(hidden_company).env
        hidden_journal = self._ensure_journal(
            hidden_env, hidden_company, 'general', 'HGL', 'Hidden GL')
        hidden_cash = hidden_env['account.account'].create({
            'code': 'H1000', 'name': 'Hidden Cash',
            'account_type': 'asset_cash',
            'company_ids': [(6, 0, [hidden_company.id])],
        })
        hidden_income = hidden_env['account.account'].create({
            'code': 'H4000', 'name': 'Hidden Revenue',
            'account_type': 'income',
            'company_ids': [(6, 0, [hidden_company.id])],
        })
        hidden_move = hidden_env['account.move'].create({
            'move_type': 'entry', 'date': '2026-06-15',
            'journal_id': hidden_journal.id,
            'line_ids': [
                (0, 0, {'name': 'Hidden', 'account_id': hidden_cash.id,
                        'debit': 999.0}),
                (0, 0, {'name': 'Hidden', 'account_id': hidden_income.id,
                        'credit': 999.0}),
            ],
        })
        hidden_move.action_post()
        self._post_in_period([
            {'account': self.account_cash, 'debit': 25.0},
            {'account': self.account_revenue, 'credit': 25.0},
        ])
        restricted = new_test_user(
            self.env,
            login='eh_gl_company_isolation',
            groups='eh_account_base.group_eh_user',
            company_id=self.company.id,
            company_ids=[(6, 0, [self.company.id])],
        )
        restricted_report = self.report.with_user(restricted)
        restricted_handler = self.handler.with_user(restricted)
        with patch.object(
            type(restricted_handler), 'compute',
            wraps=restricted_handler.compute,
        ) as compute:
            with self.assertRaises(AccessError):
                restricted_report.render(dict(
                    self.options,
                    company_ids=[self.company.id, hidden_company.id],
                ), use_cache=False)
            # The orchestrator rejects the unauthorised scope before any
            # ledger SQL is dispatched, so the hidden company's 999 can
            # neither leak nor influence a total.
            compute.assert_not_called()

    def test_foreign_amount_uses_aml_currency_precision(self):
        kwd = self.env.ref('base.KWD')
        line = self.handler._entry_line(
            {
                'aml_id': 1,
                'account_id': self.account_cash.id,
                'date': fields.Date.from_string('2026-06-15'),
                'currency_id': kwd.id,
                'amount_currency': 0.001,
            },
            debit=0.0,
            credit=0.0,
            running_balance=0.0,
            foreign_currencies={kwd.id: kwd},
        )
        self.assertEqual(
            self._column_value(line, 'foreign'),
            "0.001 %s" % kwd.name,
        )

    # ---- core rendering ----

    def test_kwd_precision_keeps_milliunit(self):
        kwd = self.env.ref('base.KWD')
        self.assertEqual(kwd.decimal_places, 3)
        options = dict(
            self.options,
            presentation_currency_id=kwd.id,
        )
        entries = [{
            'aml_id': 1,
            'account_id': self.account_cash.id,
            'account_code': self.account_cash.code,
            'account_name': self.account_cash.name,
            'date': fields.Date.from_string('2026-06-15'),
            'journal_code': self.journal_misc.code,
            'move_name': 'KWD/0001',
            'partner_name': '',
            'line_label': 'KWD milliunit',
            'ref': '',
            'debit': 0.001,
            'credit': 0.0,
            'currency_id': False,
            'amount_currency': 0.0,
        }]
        lines = self.handler._build_lines(
            {}, entries, show_zero=False,
            options=options,
            presentation_converted=True,
        )
        aml_line = self._line_by_kind(lines, 'aml')
        total_line = self._line_by_kind(lines, 'account_total')
        self.assertEqual(self._column_value(aml_line, 'debit'), 0.001)
        self.assertEqual(self._column_value(aml_line, 'balance'), 0.001)
        self.assertEqual(self._column_value(total_line, 'balance'), 0.001)

    def test_account_header_opening_entry_total_present(self):
        self._post_in_period([
            {'account': self.account_revenue, 'credit': 1000.0,
             'partner': self.partner_a},
            {'account': self.account_cash, 'debit': 1000.0},
        ])
        result = self.handler.compute(self.options)
        cash_lines = self._lines_for_account(result, self.account_cash.id)
        # Each account contributes header + opening + entry + total => 4 lines
        # at minimum (header counts itself among lines for the account).
        kinds = [(l.get('meta') or {}).get('kind') for l in cash_lines]
        self.assertIn('account_header', kinds)
        self.assertIn('opening_balance', kinds)
        self.assertIn('aml', kinds)
        self.assertIn('account_total', kinds)

    def test_running_balance_increments_correctly(self):
        # Two postings on the cash account.
        self._post_in_period([
            {'account': self.account_revenue, 'credit': 100.0},
            {'account': self.account_cash, 'debit': 100.0},
        ], date_str='2026-06-15')
        self._post_in_period([
            {'account': self.account_expense, 'debit': 30.0},
            {'account': self.account_cash, 'credit': 30.0},
        ], date_str='2026-06-20')

        result = self.handler.compute(self.options)
        cash_lines = self._lines_for_account(result, self.account_cash.id)
        # Sort by their position in the result for the cash account.
        aml_lines = [
            l for l in cash_lines
            if (l.get('meta') or {}).get('kind') == 'aml'
        ]
        self.assertEqual(len(aml_lines), 2)
        balances = [self._column_value(l, 'balance') for l in aml_lines]
        # Opening = 0 (no prior). After +100 -> 100. After -30 -> 70.
        self.assertAlmostEqual(balances[0], 100.0, places=2)
        self.assertAlmostEqual(balances[1], 70.0, places=2)

    def test_opening_balance_from_prior_period(self):
        self.post_balanced_move(
            [
                {'account': self.account_revenue, 'credit': 500.0},
                {'account': self.account_cash, 'debit': 500.0},
            ],
            date=fields.Date.from_string('2025-12-15'),
        )
        result = self.handler.compute(self.options)
        cash_lines = self._lines_for_account(result, self.account_cash.id)
        opening_line = self._line_by_kind(cash_lines, 'opening_balance')
        self.assertIsNotNone(opening_line)
        self.assertAlmostEqual(
            self._column_value(opening_line, 'balance'), 500.0, places=2,
        )

    def test_closing_balance_on_total_line(self):
        self._post_in_period([
            {'account': self.account_revenue, 'credit': 250.0},
            {'account': self.account_cash, 'debit': 250.0},
        ])
        self._post_in_period([
            {'account': self.account_expense, 'debit': 50.0},
            {'account': self.account_cash, 'credit': 50.0},
        ])
        result = self.handler.compute(self.options)
        cash_lines = self._lines_for_account(result, self.account_cash.id)
        total_line = self._line_by_kind(cash_lines, 'account_total')
        self.assertIsNotNone(total_line)
        # Closing for cash: 0 + 250 - 50 = 200.
        self.assertAlmostEqual(
            self._column_value(total_line, 'balance'), 200.0, places=2,
        )

    def test_entry_line_columns_populated(self):
        move = self._post_in_period([
            {'account': self.account_revenue, 'credit': 75.0,
             'name': 'Sales of widgets', 'partner': self.partner_a},
            {'account': self.account_cash, 'debit': 75.0},
        ])
        result = self.handler.compute(self.options)
        cash_lines = self._lines_for_account(result, self.account_cash.id)
        aml_line = self._line_by_kind(cash_lines, 'aml')
        self.assertIsNotNone(aml_line)
        self.assertEqual(self._column_value(aml_line, 'journal'),
                         self.journal_misc.code)
        self.assertEqual(self._column_value(aml_line, 'move'), move.name)
        self.assertEqual(self._column_value(aml_line, 'date'), '2026-06-15')

    # ---- date filtering ----

    def test_out_of_period_entries_not_listed_as_aml(self):
        self.post_balanced_move(
            [
                {'account': self.account_revenue, 'credit': 1000.0},
                {'account': self.account_cash, 'debit': 1000.0},
            ],
            date=fields.Date.from_string('2025-12-15'),
        )
        self.post_balanced_move(
            [
                {'account': self.account_revenue, 'credit': 999.0},
                {'account': self.account_cash, 'debit': 999.0},
            ],
            date=fields.Date.from_string('2027-02-01'),
        )
        result = self.handler.compute(self.options)
        cash_lines = self._lines_for_account(result, self.account_cash.id)
        aml_lines = [
            l for l in cash_lines
            if (l.get('meta') or {}).get('kind') == 'aml'
        ]
        # The 2025 entry contributes to opening; 2027 entry is excluded.
        # No aml rows should appear in the period.
        self.assertEqual(len(aml_lines), 0)
        # Opening reflects the 2025 entry only.
        opening = self._line_by_kind(cash_lines, 'opening_balance')
        self.assertAlmostEqual(
            self._column_value(opening, 'balance'), 1000.0, places=2,
        )

    # ---- state filtering ----

    def test_posted_only_excludes_draft(self):
        self.env['account.move'].create({
            'move_type': 'entry',
            'journal_id': self.journal_misc.id,
            'date': '2026-06-15',
            'line_ids': [
                (0, 0, {'account_id': self.account_revenue.id, 'credit': 999.0}),
                (0, 0, {'account_id': self.account_cash.id, 'debit': 999.0}),
            ],
        })
        result = self.handler.compute(self.options)
        cash_lines = self._lines_for_account(result, self.account_cash.id)
        aml_lines = [
            l for l in cash_lines
            if (l.get('meta') or {}).get('kind') == 'aml'
        ]
        self.assertEqual(len(aml_lines), 0)

    def test_posted_only_false_includes_draft(self):
        self.env['account.move'].create({
            'move_type': 'entry',
            'journal_id': self.journal_misc.id,
            'date': '2026-06-15',
            'line_ids': [
                (0, 0, {'account_id': self.account_revenue.id, 'credit': 333.0}),
                (0, 0, {'account_id': self.account_cash.id, 'debit': 333.0}),
            ],
        })
        opts = dict(self.options)
        opts['posted_only'] = False
        result = self.handler.compute(opts)
        cash_lines = self._lines_for_account(result, self.account_cash.id)
        aml_lines = [
            l for l in cash_lines
            if (l.get('meta') or {}).get('kind') == 'aml'
        ]
        self.assertEqual(len(aml_lines), 1)

    def test_cancelled_excluded(self):
        move = self._post_in_period([
            {'account': self.account_revenue, 'credit': 444.0},
            {'account': self.account_cash, 'debit': 444.0},
        ])
        move.button_cancel()
        result = self.handler.compute(self.options)
        cash_lines = self._lines_for_account(result, self.account_cash.id)
        # No entries should be listed for cash; the account header may not
        # appear at all if there is nothing else.
        aml_lines = [
            l for l in cash_lines
            if (l.get('meta') or {}).get('kind') == 'aml'
        ]
        self.assertEqual(len(aml_lines), 0)

    # ---- filter narrowing ----

    def test_account_filter(self):
        self._post_in_period([
            {'account': self.account_revenue, 'credit': 100.0},
            {'account': self.account_cash, 'debit': 100.0},
        ])
        opts = dict(self.options)
        opts['account_ids'] = [self.account_revenue.id]
        result = self.handler.compute(opts)
        # Only revenue lines and an account header should be present.
        rev_lines = self._lines_for_account(result, self.account_revenue.id)
        cash_lines = self._lines_for_account(result, self.account_cash.id)
        self.assertGreater(len(rev_lines), 0)
        self.assertEqual(len(cash_lines), 0)

    def test_partner_filter(self):
        self._post_in_period([
            {'account': self.account_revenue, 'credit': 100.0,
             'partner': self.partner_a},
            {'account': self.account_cash, 'debit': 100.0},
        ])
        self._post_in_period([
            {'account': self.account_revenue, 'credit': 200.0,
             'partner': self.partner_b},
            {'account': self.account_cash, 'debit': 200.0},
        ])
        opts = dict(self.options)
        opts['partner_ids'] = [self.partner_a.id]
        result = self.handler.compute(opts)
        rev_lines = self._lines_for_account(result, self.account_revenue.id)
        aml_lines = [
            l for l in rev_lines
            if (l.get('meta') or {}).get('kind') == 'aml'
        ]
        self.assertEqual(len(aml_lines), 1)

    # ---- error handling ----

    def test_missing_date_from_raises(self):
        bad = dict(self.options)
        bad['date'] = {'date_to': '2026-12-31'}
        with self.assertRaises(UserError):
            self.handler.compute(bad)

    def test_missing_date_to_raises(self):
        bad = dict(self.options)
        bad['date'] = {'date_from': '2026-01-01'}
        with self.assertRaises(UserError):
            self.handler.compute(bad)

    def test_prefix_sql_failure_rolls_back_and_never_resets_running(self):
        class BrokenPrefixQuery:

            @staticmethod
            def _debit_expr():
                return SQL("0")

            @staticmethod
            def _credit_expr():
                return SQL("0")

            def select(self, *args):
                return self

            def order_by(self, *args):
                return self

            def limit(self, *args):
                return self

            @staticmethod
            def build():
                return SQL(
                    "SELECT (1 / 0)::numeric AS debit, "
                    "0::numeric AS credit"
                )

        HandlerClass = type(self.handler)
        with patch.object(
            HandlerClass,
            '_expand_build_page_query',
            return_value=BrokenPrefixQuery(),
        ):
            with self.assertRaisesRegex(UserError, 'running balance'):
                self.handler._expand_prefix_balance(
                    self.options,
                    self.account_cash.id,
                    fields.Date.from_string('2026-01-01'),
                    fields.Date.from_string('2026-12-31'),
                    True,
                    [self.company.id],
                    1,
                    currency=self.company.currency_id,
                )

        # Forced SQL error must be rolled back before control returns.
        self.env.cr.execute(SQL("SELECT 21 + 21"))
        self.assertEqual(self.env.cr.fetchone()[0], 42)

    # ---- orchestrator wiring ----

    def test_orchestrator_renders(self):
        self._post_in_period([
            {'account': self.account_revenue, 'credit': 100.0},
            {'account': self.account_cash, 'debit': 100.0},
        ])
        result = self.report.render(self.options)
        self.assertFalse(result['from_cache'])
        self.assertIn('execution_id', result)
        self.assertGreater(len(result['lines']), 0)

    def test_orchestrator_cache_hit_on_second_render(self):
        self._post_in_period([
            {'account': self.account_revenue, 'credit': 100.0},
            {'account': self.account_cash, 'debit': 100.0},
        ])
        first = self.report.render(self.options)
        second = self.report.render(self.options)
        self.assertFalse(first['from_cache'])
        self.assertTrue(second['from_cache'])
        self.assertEqual(len(first['lines']), len(second['lines']))

    # ---- drill down ----

    def test_drilldown_aml_line_opens_move_form(self):
        move = self._post_in_period([
            {'account': self.account_revenue, 'credit': 75.0},
            {'account': self.account_cash, 'debit': 75.0},
        ])
        # Find any aml line on the cash account.
        result = self.handler.compute(self.options)
        cash_lines = self._lines_for_account(result, self.account_cash.id)
        aml_line = self._line_by_kind(cash_lines, 'aml')
        self.assertIsNotNone(aml_line)
        action = self.handler.get_drilldown_action(self.options, aml_line['id'])
        self.assertIsNotNone(action)
        self.assertEqual(action['res_model'], 'account.move')
        self.assertEqual(action['res_id'], move.id)

    def test_drilldown_account_header_uses_base_default(self):
        self._post_in_period([
            {'account': self.account_revenue, 'credit': 75.0},
            {'account': self.account_cash, 'debit': 75.0},
        ])
        action = self.handler.get_drilldown_action(
            self.options, "account-%s" % self.account_revenue.id,
        )
        self.assertIsNotNone(action)
        self.assertEqual(action['res_model'], 'account.move.line')

    def test_drilldown_opening_or_total_returns_none(self):
        self.assertIsNone(self.handler.get_drilldown_action(
            self.options, "account-%s-opening" % self.account_cash.id,
        ))
        self.assertIsNone(self.handler.get_drilldown_action(
            self.options, "account-%s-total" % self.account_cash.id,
        ))

    def test_drilldown_aml_with_unknown_id_returns_none(self):
        self.assertIsNone(self.handler.get_drilldown_action(
            self.options, "aml-99999999",
        ))

    # ---- XLSX export ----

    def test_xlsx_export_renders_workbook(self):
        self._post_in_period([
            {'account': self.account_revenue, 'credit': 100.0,
             'partner': self.partner_a},
            {'account': self.account_cash, 'debit': 100.0},
        ])
        content = self.report.render_xlsx(self.options)
        self.assertEqual(content[:2], b'PK')
        self.assertGreater(len(content), 1000,
                           "XLSX should contain meaningful content")


@tagged('eh_account_dynamic_reports', 'integration', 'post_install',
        '-at_install')
class TestGeneralLedgerFiscalYearWS4(EhAccountIntegrationTestCase):
    """WS4: GL fiscal-year opening + unaffected-earnings reclassification."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.handler = cls.env[
            'eh.account.dynamic.report.handler.general_ledger']
        # Demo-less Odoo 16 has no pre-provisioned unaffected-earnings
        # account. Seed that chart prerequisite explicitly so fiscal-year
        # reset/reclassification tests exercise the configured path; the
        # report itself remains read-only when this account is absent.
        cls.unaffected_account = (
            cls.company.get_unaffected_earnings_account()
        )

    def _opts(self, date_from='2026-01-01', date_to='2026-12-31', **kw):
        opts = {
            'date': {'date_from': date_from, 'date_to': date_to},
            'company_ids': [self.company.id],
            'posted_only': True, 'show_zero': False,
        }
        opts.update(kw)
        return opts

    def test_opening_excludes_prior_year_pl(self):
        # Prior-year revenue 700 (credit) before FY start. The revenue
        # account's opening must exclude it (P&L resets at FY boundary).
        self.post_balanced_move(
            [{'account': self.account_revenue, 'credit': 700.0},
             {'account': self.account_cash, 'debit': 700.0}],
            date=fields.Date.from_string('2025-09-15'))
        opening = self.handler._fetch_opening_balances(
            company_ids=[self.company.id],
            date_from=fields.Date.from_string('2026-01-01'),
            posted_only=True, options=self._opts(),
        )
        # Revenue (income) opening is excluded -> not present or zero.
        self.assertAlmostEqual(
            opening.get(self.account_revenue.id, 0.0), 0.0, places=2)
        # Cash (balance sheet) keeps its 700 opening.
        self.assertAlmostEqual(
            opening.get(self.account_cash.id, 0.0), 700.0, places=2)

    def test_unaffected_earnings_opening_equals_prior_pl(self):
        # The unaffected-earnings account opening equals net prior-year P&L.
        self.post_balanced_move(
            [{'account': self.account_revenue, 'credit': 1000.0},
             {'account': self.account_cash, 'debit': 1000.0}],
            date=fields.Date.from_string('2025-06-15'))
        self.post_balanced_move(
            [{'account': self.account_expense, 'debit': 400.0},
             {'account': self.account_cash, 'credit': 400.0}],
            date=fields.Date.from_string('2025-07-15'))
        opening = self.handler._fetch_opening_balances(
            company_ids=[self.company.id],
            date_from=fields.Date.from_string('2026-01-01'),
            posted_only=True, options=self._opts(),
        )
        unaffected = self.unaffected_account
        # Net prior P&L balance: revenue -1000 (credit) + expense +400 (debit)
        # = -600 (net credit = 600 profit), surfaced as the unaffected opening.
        self.assertAlmostEqual(
            opening.get(unaffected.id, 0.0), -600.0, places=2)

    def test_branch_fy_uses_root_owned_unaffected_account(self):
        if 'company_ids' not in self.env['account.account']._fields:
            self.skipTest('shared root-owned accounts start in Odoo 18')

        branch = self.env['res.company'].create({
            'name': 'GL Shared-chart Branch',
            'parent_id': self.company.id,
        })
        self.env.user.write({'company_ids': [(4, branch.id)]})
        branch_env = self.env['account.move'].with_context(
            allowed_company_ids=[self.company.id, branch.id],
        ).with_company(branch).env
        branch_journal = self._ensure_journal(
            branch_env, branch, 'general', 'GLBR', 'GL Branch Journal',
        )
        move = branch_env['account.move'].create({
            'move_type': 'entry',
            'journal_id': branch_journal.id,
            'date': fields.Date.from_string('2025-09-15'),
            'line_ids': [
                (0, 0, {
                    'name': 'Branch prior-year revenue',
                    'account_id': self.account_revenue.id,
                    'credit': 700.0,
                }),
                (0, 0, {
                    'name': 'Branch prior-year cash',
                    'account_id': self.account_cash.id,
                    'debit': 700.0,
                }),
            ],
        })
        move.action_post()

        handler = branch_env[
            'eh.account.dynamic.report.handler.general_ledger'
        ]
        options = {
            'date': {'date_from': '2026-01-01', 'date_to': '2026-12-31'},
            'company_ids': [branch.id],
            'posted_only': True,
            'show_zero': False,
        }
        opening = handler._fetch_opening_balances(
            company_ids=[branch.id],
            date_from=fields.Date.from_string('2026-01-01'),
            posted_only=True,
            options=options,
        )

        self.assertEqual(self.unaffected_account.company_ids, self.company)
        self.assertEqual(
            handler._unaffected_account_ids([branch.id]),
            {branch.id: self.unaffected_account.id},
        )
        self.assertAlmostEqual(
            opening.get(self.account_revenue.id, 0.0), 0.0, places=2,
        )
        self.assertAlmostEqual(
            opening.get(self.account_cash.id, 0.0), 700.0, places=2,
        )
        self.assertAlmostEqual(
            opening.get(self.unaffected_account.id, 0.0), -700.0, places=2,
        )
        self.assertAlmostEqual(sum(opening.values()), 0.0, places=2)

    def test_opening_parity_sum_equals_all_prior_balance(self):
        # Sum of all account openings (incl unaffected) must equal the sum of
        # all aml.balance before date_from: the roll reclassifies, never loses.
        self.post_balanced_move(
            [{'account': self.account_revenue, 'credit': 900.0},
             {'account': self.account_cash, 'debit': 900.0}],
            date=fields.Date.from_string('2025-05-15'))
        self.post_balanced_move(
            [{'account': self.account_expense, 'debit': 250.0},
             {'account': self.account_cash, 'credit': 250.0}],
            date=fields.Date.from_string('2025-08-15'))
        opening = self.handler._fetch_opening_balances(
            company_ids=[self.company.id],
            date_from=fields.Date.from_string('2026-01-01'),
            posted_only=True, options=self._opts(),
        )
        # Every prior move is balanced, so the signed sum of all openings is 0.
        self.assertAlmostEqual(sum(opening.values()), 0.0, places=2)

    def test_fy_source_account_filters_do_not_expose_destination(self):
        self.post_balanced_move(
            [{'account': self.account_revenue, 'credit': 700.0},
             {'account': self.account_cash, 'debit': 700.0}],
            date=fields.Date.from_string('2025-09-15'),
        )

        HandlerClass = type(self.handler)
        with patch.object(
            HandlerClass,
            '_fetch_prior_pl_roll',
            wraps=self.handler._fetch_prior_pl_roll,
        ) as source_roll:
            by_account = self.handler._fetch_opening_balances(
                company_ids=[self.company.id],
                date_from=fields.Date.from_string('2026-01-01'),
                posted_only=True,
                options=self._opts(account_ids=[self.account_revenue.id]),
            )
            by_type = self.handler._fetch_opening_balances(
                company_ids=[self.company.id],
                date_from=fields.Date.from_string('2026-01-01'),
                posted_only=True,
                options=self._opts(account_type_ids=['income']),
            )

        self.assertFalse(source_roll.called)
        self.assertNotIn(self.unaffected_account.id, by_account)
        self.assertNotIn(self.unaffected_account.id, by_type)

    def test_fy_destination_filters_keep_full_prior_pl_source(self):
        self.post_balanced_move(
            [{'account': self.account_revenue, 'credit': 1000.0},
             {'account': self.account_cash, 'debit': 1000.0}],
            date=fields.Date.from_string('2025-06-15'),
        )
        self.post_balanced_move(
            [{'account': self.account_expense, 'debit': 400.0},
             {'account': self.account_cash, 'credit': 400.0}],
            date=fields.Date.from_string('2025-07-15'),
        )

        by_account = self.handler._fetch_opening_balances(
            company_ids=[self.company.id],
            date_from=fields.Date.from_string('2026-01-01'),
            posted_only=True,
            options=self._opts(
                account_ids=[self.unaffected_account.id],
            ),
        )
        self.assertAlmostEqual(
            by_account.get(self.unaffected_account.id, 0.0),
            -600.0,
            places=2,
        )

        by_type = self.handler._fetch_opening_balances(
            company_ids=[self.company.id],
            date_from=fields.Date.from_string('2026-01-01'),
            posted_only=True,
            options=self._opts(account_type_ids=['equity_unaffected']),
        )
        self.assertAlmostEqual(
            by_type.get(self.unaffected_account.id, 0.0),
            -600.0,
            places=2,
        )

    def test_report_does_not_create_missing_unaffected_account(self):
        """Rendering is read-only even when chart setup is incomplete."""
        company = self.env['res.company'].create({
            'name': 'GL Read-only Missing Unaffected Account',
        })
        Account = self.env['account.account'].with_company(company)
        domain = [('account_type', '=', 'equity_unaffected')]
        if 'company_ids' in Account._fields:
            domain.append(('company_ids', 'in', [company.id]))
        else:
            domain.append(('company_id', '=', company.id))
        self.assertFalse(Account.search(domain))
        self.assertEqual(
            self.handler._unaffected_account_ids([company.id]),
            {},
        )
        self.assertFalse(Account.search(domain))

    def test_missing_unaffected_account_preserves_prior_pl_opening(self):
        self.post_balanced_move(
            [{'account': self.account_revenue, 'credit': 700.0},
             {'account': self.account_cash, 'debit': 700.0}],
            date=fields.Date.from_string('2025-09-15'),
        )
        HandlerClass = type(self.handler)
        with patch.object(
            HandlerClass, '_unaffected_account_ids', return_value={},
        ):
            opening = self.handler._fetch_opening_balances(
                company_ids=[self.company.id],
                date_from=fields.Date.from_string('2026-01-01'),
                posted_only=True,
                options=self._opts(),
            )
        self.assertAlmostEqual(
            opening.get(self.account_revenue.id, 0.0), -700.0, places=2,
        )
        self.assertAlmostEqual(
            opening.get(self.account_cash.id, 0.0), 700.0, places=2,
        )
        self.assertAlmostEqual(sum(opening.values()), 0.0, places=2)

    def test_fy_sql_failure_rolls_back_before_legacy_fallback(self):
        self.post_balanced_move(
            [{'account': self.account_cash, 'debit': 325.0},
             {'account': self.account_receivable, 'credit': 325.0}],
            date=fields.Date.from_string('2025-09-15'),
        )

        def force_sql_error(*args, **kwargs):
            self.env.cr.execute(SQL("SELECT 1 / 0"))

        HandlerClass = type(self.handler)
        with patch.object(
            HandlerClass,
            '_fetch_opening_balances_fy_aware',
            side_effect=force_sql_error,
        ):
            opening = self.handler._fetch_opening_balances(
                company_ids=[self.company.id],
                date_from=fields.Date.from_string('2026-01-01'),
                posted_only=True,
                options=self._opts(),
            )

        self.assertAlmostEqual(
            opening.get(self.account_cash.id, 0.0), 325.0, places=2,
        )
        self.assertAlmostEqual(
            opening.get(self.account_receivable.id, 0.0), -325.0, places=2,
        )
        self.env.cr.execute(SQL("SELECT 20 + 22"))
        self.assertEqual(self.env.cr.fetchone()[0], 42)

    def test_balance_sheet_account_opening_unchanged(self):
        # Regression: a pure balance-sheet opening (no P&L) is byte-identical
        # to the legacy all-time-before rule.
        self.post_balanced_move(
            [{'account': self.account_cash, 'debit': 500.0},
             {'account': self.account_receivable, 'credit': 500.0}],
            date=fields.Date.from_string('2025-12-15'))
        opening = self.handler._fetch_opening_balances(
            company_ids=[self.company.id],
            date_from=fields.Date.from_string('2026-01-01'),
            posted_only=True, options=self._opts(),
        )
        self.assertAlmostEqual(
            opening.get(self.account_cash.id, 0.0), 500.0, places=2)
        self.assertAlmostEqual(
            opening.get(self.account_receivable.id, 0.0), -500.0, places=2)


@tagged('eh_account_dynamic_reports', 'integration', 'post_install',
        '-at_install')
class TestGeneralLedgerMultiCurrencyWS4(EhAccountIntegrationTestCase):
    """WS4: GL consolidated opening converts cross-company balances."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.handler = cls.env[
            'eh.account.dynamic.report.handler.general_ledger']
        cls.company_a = cls.company
        cls.base_currency = cls.company_a.currency_id
        cls.currency_b = cls.env['res.currency'].create({
            'name': 'YBT', 'symbol': 'Y',
            'rounding': cls.base_currency.rounding,
        })
        cls.env['res.currency.rate'].create({
            'currency_id': cls.currency_b.id, 'name': '2020-01-01',
            'rate': 4.0,
        })
        cls.company_b = cls.env['res.company'].create({
            'name': 'FX GL Co B', 'currency_id': cls.currency_b.id,
        })
        cls.env['res.currency.rate'].create({
            'currency_id': cls.base_currency.id,
            'company_id': cls.company_b.id,
            'name': '2020-01-01',
            # Company currency is unit rate in its own company; target rate
            # 0.5 makes one YBT minor unit half one target minor unit.
            'rate': 0.5,
        })
        cls.env.user.company_ids = [(4, cls.company_b.id)]
        cls.journal_b = cls.env['account.journal'].create({
            'name': 'Misc GLB', 'code': 'MGLB', 'type': 'general',
            'company_id': cls.company_b.id,
        })
        cls.cash_b = cls.env['account.account'].create({
            'code': '1003B', 'name': 'Cash GLB', 'account_type': 'asset_cash',
            'company_ids': [(6, 0, [cls.company_b.id])],
        })
        cls.equity_b = cls.env['account.account'].create({
            'code': '3003B', 'name': 'Equity GLB', 'account_type': 'equity',
            'company_ids': [(6, 0, [cls.company_b.id])],
        })
        cls.report = cls.env['eh.account.dynamic.report'].search(
            [('code', '=', 'general_ledger')], limit=1,
        )
        if not cls.report:
            cls.report = cls.env['eh.account.dynamic.report'].create({
                'code': 'general_ledger',
                'name': 'General Ledger',
                'handler_model':
                    'eh.account.dynamic.report.handler.general_ledger',
            })

    @staticmethod
    def _column(line, label):
        return next(
            column['value'] for column in line['columns']
            if column['expression_label'] == label
        )

    def _consolidated_options(self, **extra):
        options = {
            'date': {'date_from': '2026-01-01', 'date_to': '2026-12-31'},
            'company_ids': [self.company_a.id, self.company_b.id],
            'posted_only': True,
            'show_zero': False,
            'presentation_currency_id': self.base_currency.id,
        }
        options.update(extra)
        return options

    def test_consolidated_opening_converts_company_b(self):
        # Balance-sheet opening before date_from. A: cash 800 base. B: cash
        # 800 currency-B, which converts to base via the ORM rate.
        self.post_balanced_move(
            [{'account': self.account_cash, 'debit': 800.0},
             {'account': self.account_receivable, 'credit': 800.0}],
            date=fields.Date.from_string('2025-12-10'))
        self.post_balanced_move(
            [{'account': self.cash_b, 'debit': 800.0},
             {'account': self.equity_b, 'credit': 800.0}],
            journal=self.journal_b,
            date=fields.Date.from_string('2025-12-10'))

        options = {
            'date': {'date_from': '2026-01-01', 'date_to': '2026-12-31'},
            'company_ids': [self.company_a.id, self.company_b.id],
            'posted_only': True, 'show_zero': False,
            'presentation_currency_id': self.base_currency.id,
        }
        opening = self.handler._fetch_opening_balances(
            company_ids=[self.company_a.id, self.company_b.id],
            date_from=fields.Date.from_string('2026-01-01'),
            posted_only=True, options=options,
        )
        b_cash_in_base = self.currency_b._convert(
            800.0, self.base_currency, self.company_b,
            fields.Date.from_string('2026-12-31'))
        self.assertAlmostEqual(
            opening.get(self.account_cash.id, 0.0), 800.0, places=2)
        self.assertAlmostEqual(
            opening.get(self.cash_b.id, 0.0),
            self.base_currency.round(b_cash_in_base),
            places=2)
        # Prove conversion happened (raw would be 800 for B too).
        if not self.base_currency.is_zero(b_cash_in_base - 800.0):
            self.assertNotAlmostEqual(
                opening.get(self.cash_b.id, 0.0), 800.0, places=2)

    def test_foreign_currency_is_classified_per_aml_company(self):
        """Another scoped company's ledger currency may still be foreign."""
        move = self.env['account.move'].create({
            'move_type': 'entry',
            'journal_id': self.journal_misc.id,
            'date': fields.Date.from_string('2026-06-01'),
            'line_ids': [
                (0, 0, {
                    'account_id': self.account_cash.id,
                    'debit': 100.0,
                    'currency_id': self.currency_b.id,
                    'amount_currency': 25.0,
                    'name': 'Company A foreign YBT line',
                }),
                (0, 0, {
                    'account_id': self.account_equity.id,
                    'credit': 100.0,
                    'currency_id': self.base_currency.id,
                    'amount_currency': -100.0,
                }),
            ],
        })
        move.action_post()
        foreign_line = move.line_ids.filtered(
            lambda line: line.account_id == self.account_cash
        )
        options = {
            'date': {'date_from': '2026-01-01', 'date_to': '2026-12-31'},
            'company_ids': [self.company_a.id, self.company_b.id],
            'posted_only': True,
            'show_zero': False,
            'presentation_currency_id': self.base_currency.id,
        }

        payload = self.handler.compute(options)
        eager = next(
            line for line in payload['lines']
            if line['id'] == 'aml-%s' % foreign_line.id
        )
        eager_foreign = next(
            column['value'] for column in eager['columns']
            if column['expression_label'] == 'foreign'
        )
        self.assertIn(self.currency_b.name, eager_foreign)

        lazy = self.handler.expand_account_line(
            dict(options, lazy_expand=True),
            'account-%s' % self.account_cash.id,
            offset=0,
            limit=80,
        )
        lazy_line = next(
            line for line in lazy['child_lines']
            if line['id'] == 'aml-%s' % foreign_line.id
        )
        lazy_foreign = next(
            column['value'] for column in lazy_line['columns']
            if column['expression_label'] == 'foreign'
        )
        self.assertIn(self.currency_b.name, lazy_foreign)

    def test_half_minor_fx_foots_lazy_eager_expand_and_xlsx(self):
        """Displayed AML rounding owns totals in every GL delivery path."""
        converted = self.currency_b._convert(
            0.01,
            self.base_currency,
            self.company_b,
            fields.Date.from_string('2026-12-31'),
            round=False,
        )
        self.assertAlmostEqual(converted, 0.005, places=6)

        labels = ['Half-minor FX 1', 'Half-minor FX 2']
        for index, label in enumerate(labels, start=1):
            self.post_balanced_move(
                [
                    {'account': self.cash_b, 'debit': 0.01, 'name': label},
                    {
                        'account': self.equity_b,
                        'credit': 0.01,
                        'name': 'Half-minor offset %s' % index,
                    },
                ],
                journal=self.journal_b,
                date=fields.Date.from_string('2026-06-%02d' % index),
            )

        options = self._consolidated_options()
        eager = self.handler.compute(options)
        eager_lines = [
            line for line in eager['lines']
            if (line.get('meta') or {}).get('account_id') == self.cash_b.id
        ]
        eager_entries = [
            line for line in eager_lines
            if (line.get('meta') or {}).get('kind') == 'aml'
        ]
        eager_total = next(
            line for line in eager_lines
            if (line.get('meta') or {}).get('kind') == 'account_total'
        )
        self.assertEqual(
            [self._column(line, 'debit') for line in eager_entries],
            [0.01, 0.01],
        )
        self.assertEqual(self._column(eager_total, 'balance'), 0.02)

        lazy_options = self._consolidated_options(lazy_expand=True)
        lazy = self.handler.compute(lazy_options)
        lazy_total = next(
            line for line in lazy['lines']
            if line['id'] == 'account-%s-total' % self.cash_b.id
        )
        self.assertEqual(self._column(lazy_total, 'balance'), 0.02)

        page_1 = self.handler.expand_account_line(
            lazy_options,
            'account-%s' % self.cash_b.id,
            offset=0,
            limit=1,
        )
        page_2 = self.handler.expand_account_line(
            lazy_options,
            'account-%s' % self.cash_b.id,
            offset=page_1['next_offset'],
            limit=1,
        )
        expanded = page_1['child_lines'] + page_2['child_lines']
        self.assertEqual(
            [self._column(line, 'debit') for line in expanded],
            [0.01, 0.01],
        )
        self.assertEqual(
            [self._column(line, 'balance') for line in expanded],
            [0.01, 0.02],
        )
        self.assertEqual(self._column(expanded[-1], 'balance'), 0.02)

        workbook = load_workbook(
            BytesIO(self.report.render_xlsx(lazy_options, use_cache=False)),
            data_only=True,
        )
        rows = list(workbook.active.iter_rows(values_only=True))
        header = next(row for row in rows if row[0] == 'Description')
        debit_index = header.index('Debit')
        balance_index = header.index('Balance')
        export_entries = [row for row in rows if row[0] in labels]
        cash_b_code = self.cash_b.with_company(self.company_b).code
        export_total = next(
            row for row in rows
            if row[0] and row[0].startswith('Total ')
            and cash_b_code in row[0]
        )
        self.assertEqual(
            [row[debit_index] for row in export_entries],
            [0.01, 0.01],
        )
        self.assertEqual(export_total[balance_index], 0.02)
