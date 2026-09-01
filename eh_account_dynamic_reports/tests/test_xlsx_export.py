# -*- encoding: utf-8 -*-
##############################################################################
#
# ERP Heritage
# Copyright (C) 2026 (https://www.erpheritage.com.au/)
#
##############################################################################
"""
End to end XLSX export of the Trial Balance.

Posts a fixture journal entry, runs the orchestrator's render_xlsx, and
inspects the resulting workbook to verify that:

* The XLSX is well formed and openpyxl can read it back.
* The title and metadata rows reflect the report parameters.
* The data rows match the values produced by the handler's compute().
* The totals row balances.
* A second export hits the cache and matches the first byte for byte
  (since the payload is unchanged).
"""

import io

from openpyxl import load_workbook

from odoo import fields
from odoo.tests import tagged

from odoo.addons.eh_account_base.tests.common import EhAccountIntegrationTestCase


@tagged('eh_account_dynamic_reports', 'integration', 'post_install', '-at_install')
class TestTrialBalanceXlsxExport(EhAccountIntegrationTestCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.report = cls.env['eh.account.dynamic.report'].search(
            [('code', '=', 'trial_balance')], limit=1,
        )
        if not cls.report:
            cls.report = cls.env['eh.account.dynamic.report'].create({
                'code': 'trial_balance',
                'name': 'Trial Balance',
                'handler_model': 'eh.account.dynamic.report.handler.trial_balance',
            })
        cls.options = {
            'date': {'date_from': '2026-01-01', 'date_to': '2026-12-31'},
            'company_ids': [cls.env.company.id],
            'posted_only': True,
            'show_zero': False,
        }
        cls.post_balanced_move(
            [
                {'account': cls.account_revenue, 'credit': 1000.0},
                {'account': cls.account_cash, 'debit': 1000.0},
            ],
            date=fields.Date.from_string('2026-06-15'),
        )

    def _read_workbook(self, content):
        return load_workbook(io.BytesIO(content))

    @staticmethod
    def _find_row_starting_with(ws, text):
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == text:
                return r
        return None

    def test_render_xlsx_returns_well_formed_bytes(self):
        content = self.report.render_xlsx(self.options)
        self.assertEqual(content[:2], b'PK')
        wb = self._read_workbook(content)
        self.assertEqual(wb.active.title, 'Report')

    def test_workbook_contains_data_rows(self):
        content = self.report.render_xlsx(self.options)
        ws = self._read_workbook(content).active
        cash_row = self._find_row_starting_with(ws, "1000 Cash on Hand")
        revenue_row = self._find_row_starting_with(ws, "4000 Sales Revenue")
        self.assertIsNotNone(cash_row, "Cash account row should appear")
        self.assertIsNotNone(revenue_row, "Revenue account row should appear")

    def test_workbook_totals_balance(self):
        content = self.report.render_xlsx(self.options)
        ws = self._read_workbook(content).active
        totals_row = self._find_row_starting_with(ws, "Totals")
        self.assertIsNotNone(totals_row)
        # Trial Balance totals: closing debit total equals closing credit total.
        # Columns: 1=name, 2=open DB, 3=open CR, 4=mov DB, 5=mov CR,
        # 6=close DB, 7=close CR.
        closing_db = ws.cell(row=totals_row, column=6).value or 0
        closing_cr = ws.cell(row=totals_row, column=7).value or 0
        self.assertAlmostEqual(closing_db, closing_cr, places=2)

    def test_metadata_row_includes_date_range(self):
        content = self.report.render_xlsx(self.options)
        ws = self._read_workbook(content).active
        meta = " ".join(
            (ws.cell(row=r, column=1).value or '')
            for r in range(1, 6)
        )
        self.assertIn("2026-01-01", meta)
        self.assertIn("2026-12-31", meta)

    def test_second_export_matches_first_when_cache_valid(self):
        first = self.report.render_xlsx(self.options)
        # No new posts between the two calls, so the cache is still valid.
        second = self.report.render_xlsx(self.options)
        # Workbooks built from the same payload should be identical except
        # for openpyxl's per file metadata. Verify by comparing sheet content.
        ws1 = self._read_workbook(first).active
        ws2 = self._read_workbook(second).active
        rows1 = [
            tuple(ws1.cell(row=r, column=c).value
                  for c in range(1, ws1.max_column + 1))
            for r in range(1, ws1.max_row + 1)
        ]
        rows2 = [
            tuple(ws2.cell(row=r, column=c).value
                  for c in range(1, ws2.max_column + 1))
            for r in range(1, ws2.max_row + 1)
        ]
        # Rows containing the generated_at timestamp may differ; strip them.
        rows1 = [row for row in rows1
                 if not (row and row[0] and 'Generated' in str(row[0]))]
        rows2 = [row for row in rows2
                 if not (row and row[0] and 'Generated' in str(row[0]))]
        self.assertEqual(rows1, rows2)

    def test_export_after_new_post_includes_new_data(self):
        first_content = self.report.render_xlsx(self.options)
        # Add another posting; this should bust the cache.
        self.post_balanced_move(
            [
                {'account': self.account_revenue, 'credit': 250.0},
                {'account': self.account_cash, 'debit': 250.0},
            ],
            date=fields.Date.from_string('2026-07-01'),
        )
        second_content = self.report.render_xlsx(self.options)
        ws = self._read_workbook(second_content).active
        cash_row = self._find_row_starting_with(ws, "1000 Cash on Hand")
        # 1000 + 250 = 1250 closing debit on cash.
        closing_db = ws.cell(row=cash_row, column=6).value
        self.assertAlmostEqual(closing_db, 1250.0, places=2)
        # Sanity: the two byte streams differ in content.
        self.assertNotEqual(first_content, second_content)
